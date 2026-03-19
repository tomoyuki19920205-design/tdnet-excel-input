import openpyxl
from pathlib import Path

p = Path(r"C:\Users\takuy\OneDrive\viewer.xlsx")
wb = openpyxl.load_workbook(p, data_only=False, keep_links=True)

print("sheets:", wb.sheetnames)

# よくあるキーシート名（存在チェック）
candidates = ["AUTO","_DATA","PL_VIEW","_SEGMENT_DATA","SEGMENT_VIEW","COMPANY_VIEW"]
for s in candidates:
    print(f"has {s} =", (s in wb.sheetnames))

wb.close()
