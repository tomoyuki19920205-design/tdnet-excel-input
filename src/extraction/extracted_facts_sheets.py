#!/usr/bin/env python3
# ============================================================
# extracted_facts_sheets.py
# ============================================================
"""
extracted_facts テーブルから data.xlsx 用の3シートデータを構築する。

  FORECAST_REVISION — 業績予想修正 before/after/delta
  MONTHLY_DATA      — 月次開示データ
  KPI_DATA          — KPI / 決算説明資料 / 補足資料データ

利用側:
  from src.extraction.extracted_facts_sheets import write_extracted_facts_sheets
  stats = write_extracted_facts_sheets(wb, conn)
"""
from __future__ import annotations

import logging
import re
import sqlite3
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

logger = logging.getLogger("gen_data")


# ============================================================
# 定数: シート名・列定義
# ============================================================

FORECAST_SHEET = "FORECAST_REVISION"
FORECAST_COLUMNS = [
    "ticker", "pubdate", "title", "doc_type",
    "period", "quarter", "metric_name",
    "before_value", "after_value", "delta_value", "delta_pct",
    "unit", "source_type", "confidence", "document_id",
    "raw_label", "normalized_label",
]

MONTHLY_SHEET = "MONTHLY_DATA"
MONTHLY_COLUMNS = [
    "ticker", "pubdate", "title", "doc_type",
    "year_month", "metric_name", "metric_value",
    "unit", "segment_name", "source_type", "confidence",
    "document_id", "raw_label", "normalized_label",
]

KPI_SHEET = "KPI_DATA"
KPI_COLUMNS = [
    "ticker", "pubdate", "title", "doc_type",
    "period", "quarter", "metric_name", "metric_value",
    "unit", "segment_name", "source_type", "confidence",
    "document_id", "table_title", "page_no",
    "raw_label", "normalized_label",
]

# FORECAST_REVISION 対象メトリック
_FORECAST_METRICS = {
    "sales", "operating_profit", "ordinary_profit", "net_income", "eps",
}

# MONTHLY_DATA 対象メトリック
_MONTHLY_METRICS = {
    "monthly_sales", "same_store_sales_yoy", "customer_count",
    "average_spend", "store_count", "utilization_rate",
    "orders", "order_backlog",
}

# KPI_DATA 対象 doc_type
_KPI_DOC_TYPES = {"presentation", "supplement", "kpi", "segment"}

# KPI_DATA 除外メトリック (forecast_revision 専用)
_FORECAST_ONLY_METRICS = {
    "forecast_sales", "forecast_operating_profit",
}

# before/after 判定パターン
_BEFORE_PATTERNS = re.compile(r"(前回予想|修正前|従来予想|前回発表)")
_AFTER_PATTERNS = re.compile(r"(今回予想|修正後|今回修正|今回発表)")


# ============================================================
# FORECAST_REVISION
# ============================================================

def build_forecast_revision_rows(conn: sqlite3.Connection) -> list[dict]:
    """
    doc_type='forecast_revision' の extracted_facts からペアリングして
    before/after/delta 行を構築する。

    ペアリングルール:
    1. raw_label から before/after キーワードで判定
    2. 同一キー (document_id, metric_name, period, quarter) で
       ちょうど2件 → id小=before, id大=after
    3. 3件以上 → ペアリングせず skip
    """
    rows = conn.execute(
        """SELECT ef.id, ef.document_id, ef.ticker, ef.period, ef.quarter,
                  ef.metric_name, ef.metric_value, ef.unit, ef.source_type,
                  ef.confidence, ef.raw_label, ef.normalized_label,
                  ef.segment_name, ef.table_title,
                  d.pubdate, d.title, d.doc_type
           FROM extracted_facts ef
           JOIN documents d ON ef.document_id = d.id
           WHERE d.doc_type = 'forecast_revision'
             AND ef.metric_value IS NOT NULL
             AND ef.metric_name IS NOT NULL
             AND ef.confidence != 'low'
           ORDER BY ef.document_id, ef.metric_name, ef.period, ef.quarter, ef.id
        """,
    ).fetchall()

    col_names = [desc[0] for desc in conn.execute(
        "SELECT * FROM extracted_facts LIMIT 0"
    ).description] if rows else []

    # グルーピング: (document_id, metric_name, period, quarter) → list
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        r = {
            "id": row[0], "document_id": row[1], "ticker": row[2],
            "period": row[3], "quarter": row[4], "metric_name": row[5],
            "metric_value": row[6], "unit": row[7], "source_type": row[8],
            "confidence": row[9], "raw_label": row[10] or "",
            "normalized_label": row[11] or "", "segment_name": row[12] or "",
            "table_title": row[13] or "",
            "pubdate": row[14] or "", "title": row[15] or "",
            "doc_type": row[16] or "",
        }
        # FORECAST 対象メトリックのみ
        if r["metric_name"] not in _FORECAST_METRICS:
            continue
        key = (r["document_id"], r["metric_name"], r["period"], r["quarter"])
        groups[key].append(r)

    result: list[dict] = []
    for key, items in groups.items():
        before_val = None
        after_val = None
        base = items[0]  # 共通情報

        if len(items) > 2:
            # 3件以上 → skip (ペアリング不能)
            logger.debug(
                f"[FORECAST] skip: {len(items)} 件 key={key}"
            )
            continue

        if len(items) == 2:
            # raw_label から判定
            r0_label = items[0]["raw_label"]
            r1_label = items[1]["raw_label"]

            if _BEFORE_PATTERNS.search(r0_label) and _AFTER_PATTERNS.search(r1_label):
                before_val = items[0]["metric_value"]
                after_val = items[1]["metric_value"]
            elif _AFTER_PATTERNS.search(r0_label) and _BEFORE_PATTERNS.search(r1_label):
                before_val = items[1]["metric_value"]
                after_val = items[0]["metric_value"]
            else:
                # id 順: 小=before, 大=after
                before_val = items[0]["metric_value"]
                after_val = items[1]["metric_value"]

        elif len(items) == 1:
            # 片方のみ
            label = items[0]["raw_label"]
            if _AFTER_PATTERNS.search(label):
                after_val = items[0]["metric_value"]
            elif _BEFORE_PATTERNS.search(label):
                before_val = items[0]["metric_value"]
            else:
                # 判定不能 → after_value のみ
                after_val = items[0]["metric_value"]

        # before/after 両方 null → skip
        if before_val is None and after_val is None:
            continue

        # delta 計算
        delta_value = None
        delta_pct = None
        if before_val is not None and after_val is not None:
            delta_value = after_val - before_val
            if before_val != 0:
                delta_pct = delta_value / abs(before_val)

        result.append({
            "ticker": base["ticker"],
            "pubdate": base["pubdate"],
            "title": base["title"],
            "doc_type": base["doc_type"],
            "period": base["period"],
            "quarter": base["quarter"],
            "metric_name": base["metric_name"],
            "before_value": before_val,
            "after_value": after_val,
            "delta_value": delta_value,
            "delta_pct": delta_pct,
            "unit": base["unit"],
            "source_type": base["source_type"],
            "confidence": base["confidence"],
            "document_id": base["document_id"],
            "raw_label": base["raw_label"],
            "normalized_label": base["normalized_label"],
        })

    # ソート: ticker ASC, pubdate DESC, period DESC, metric_name ASC
    result.sort(key=lambda r: (
        r["ticker"],
        _desc(r["pubdate"]),
        _desc(r["period"]),
        r["metric_name"],
    ))

    logger.info(f"[FORECAST] {len(result)} 行構築")
    return result


# ============================================================
# MONTHLY_DATA
# ============================================================

def build_monthly_data_rows(conn: sqlite3.Connection) -> list[dict]:
    """
    doc_type='monthly' の extracted_facts から月次データ行を構築する。
    year_month が取れない行でも pubdate を保持して出力する。
    """
    rows = conn.execute(
        """SELECT ef.document_id, ef.ticker, ef.period, ef.quarter,
                  ef.metric_name, ef.metric_value, ef.unit,
                  ef.segment_name, ef.source_type, ef.confidence,
                  ef.raw_label, ef.normalized_label,
                  d.pubdate, d.title, d.doc_type
           FROM extracted_facts ef
           JOIN documents d ON ef.document_id = d.id
           WHERE d.doc_type = 'monthly'
             AND ef.metric_value IS NOT NULL
             AND ef.metric_name IS NOT NULL
             AND ef.confidence != 'low'
           ORDER BY ef.ticker, d.pubdate DESC, ef.metric_name
        """,
    ).fetchall()

    result: list[dict] = []
    for row in rows:
        period = row[2] or ""
        pubdate = row[12] or ""

        # year_month 推定: period から YYYY-MM を取る
        year_month = _extract_year_month(period)

        result.append({
            "ticker": row[1],
            "pubdate": pubdate,
            "title": row[13] or "",
            "doc_type": row[14] or "",
            "year_month": year_month,
            "metric_name": row[4],
            "metric_value": row[5],
            "unit": row[6] or "",
            "segment_name": row[7] or "",
            "source_type": row[8] or "",
            "confidence": row[9] or "medium",
            "document_id": row[0],
            "raw_label": row[10] or "",
            "normalized_label": row[11] or "",
        })

    # ソート: ticker ASC, year_month DESC, metric_name ASC
    result.sort(key=lambda r: (
        r["ticker"],
        _desc(r["year_month"] or r["pubdate"]),
        r["metric_name"],
    ))

    logger.info(f"[MONTHLY] {len(result)} 行構築")
    return result


# ============================================================
# KPI_DATA
# ============================================================

def build_kpi_data_rows(conn: sqlite3.Connection) -> list[dict]:
    """
    doc_type in (presentation, supplement, kpi, segment) の
    extracted_facts から KPI データ行を構築する。
    """
    placeholders = ",".join(["?"] * len(_KPI_DOC_TYPES))
    rows = conn.execute(
        f"""SELECT ef.document_id, ef.ticker, ef.period, ef.quarter,
                  ef.metric_name, ef.metric_value, ef.unit,
                  ef.segment_name, ef.source_type, ef.confidence,
                  ef.page_no, ef.table_title,
                  ef.raw_label, ef.normalized_label,
                  d.pubdate, d.title, d.doc_type
           FROM extracted_facts ef
           JOIN documents d ON ef.document_id = d.id
           WHERE d.doc_type IN ({placeholders})
             AND ef.metric_value IS NOT NULL
             AND ef.metric_name IS NOT NULL
             AND ef.confidence != 'low'
           ORDER BY ef.ticker, ef.period DESC, ef.quarter DESC, ef.metric_name
        """,
        tuple(_KPI_DOC_TYPES),
    ).fetchall()

    result: list[dict] = []
    for row in rows:
        metric_name = row[4]

        # forecast_revision 専用メトリックは除外
        if metric_name in _FORECAST_ONLY_METRICS:
            continue

        # doc_type = monthly は MONTHLY_DATA 側
        doc_type = row[16] or ""
        if doc_type == "monthly":
            continue

        result.append({
            "ticker": row[1],
            "pubdate": row[14] or "",
            "title": row[15] or "",
            "doc_type": doc_type,
            "period": row[2] or "",
            "quarter": row[3] or "",
            "metric_name": metric_name,
            "metric_value": row[5],
            "unit": row[6] or "",
            "segment_name": row[7] or "",
            "source_type": row[8] or "",
            "confidence": row[9] or "medium",
            "document_id": row[0],
            "table_title": row[11] or "",
            "page_no": row[10],
            "raw_label": row[12] or "",
            "normalized_label": row[13] or "",
        })

    # ソート: ticker ASC, period DESC, quarter DESC, metric_name ASC, segment_name ASC
    result.sort(key=lambda r: (
        r["ticker"],
        _desc(r["period"]),
        _desc(r["quarter"]),
        r["metric_name"],
        r["segment_name"],
    ))

    logger.info(f"[KPI] {len(result)} 行構築")
    return result


# ============================================================
# シート書き込み
# ============================================================

def write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    rows: list[dict],
    columns: list[str],
    generated_at: str = "",
) -> int:
    """
    wb にシートを作成して rows を書き込む。
    既存の同名シートがあれば削除して再生成する。

    Returns: 書き込んだ行数
    """
    # 既存シート削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    if not rows:
        logger.info(f"[SHEET] {sheet_name}: データなし — シートスキップ")
        return 0

    ws = wb.create_sheet(sheet_name)

    # A1: 生成日時
    if generated_at:
        ws.cell(row=1, column=1, value=f"generated_at: {generated_at}")

    # ヘッダー (2行目)
    bold = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = bold

    # データ行 (3行目から)
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            # ticker は文字列として保持
            if col_name == "ticker" and val is not None:
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    logger.info(f"[SHEET] {sheet_name}: {len(rows)} 行書込み")
    return len(rows)


def write_extracted_facts_sheets(
    wb: openpyxl.Workbook,
    conn: sqlite3.Connection,
    generated_at: str = "",
) -> dict:
    """
    wb に FORECAST_REVISION / MONTHLY_DATA / KPI_DATA の3シートを書き込む。
    既存の同名シートがあれば削除して再生成する。

    Returns: {"forecast_rows": int, "monthly_rows": int, "kpi_rows": int}
    """
    # documents / extracted_facts テーブル存在チェック
    tables = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()}
    if "documents" not in tables or "extracted_facts" not in tables:
        logger.info("[EF] documents/extracted_facts テーブルなし — スキップ")
        return {"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0}

    forecast_rows = build_forecast_revision_rows(conn)
    monthly_rows = build_monthly_data_rows(conn)
    kpi_rows = build_kpi_data_rows(conn)

    fc = write_sheet(wb, FORECAST_SHEET, forecast_rows, FORECAST_COLUMNS, generated_at)
    mc = write_sheet(wb, MONTHLY_SHEET, monthly_rows, MONTHLY_COLUMNS, generated_at)
    kc = write_sheet(wb, KPI_SHEET, kpi_rows, KPI_COLUMNS, generated_at)

    return {"forecast_rows": fc, "monthly_rows": mc, "kpi_rows": kc}


# ============================================================
# ユーティリティ
# ============================================================

def _desc(s: str) -> str:
    """降順ソート用: 文字列を反転（空文字は最後尾に）"""
    if not s:
        return ""
    # ASCII 範囲の文字を反転して降順ソートに使う
    return "".join(chr(0xFFFF - ord(c)) if ord(c) < 0xFFFF else c for c in s)


_YM_RE = re.compile(r"(\d{4})-(\d{2})")


def _extract_year_month(period: str) -> str:
    """period 文字列から YYYY-MM を抽出する。取れなければ空文字。"""
    if not period:
        return ""
    m = _YM_RE.search(period)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return ""
