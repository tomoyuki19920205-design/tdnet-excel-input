# ============================================================
# excel_writer.py — openpyxlでの安全なExcel書き込み（最重要モジュール）
# ============================================================
#
# 安全要件:
# 1. A列コード検索 → startRow特定
# 2. M列R年度探索: [startRow, startRow + max_scan_rows] 厳守（150行制限）
# 3. N列四半期探索: [termRow - q_search_up, termRow + q_search_down] 近傍限定
# 4. 競合検知: 空→書き込み / 同値→スキップ / 異値→conflict_detected
# 5. 保存リトライ: 5回（3,5,8,13,21秒）
# ============================================================
from __future__ import annotations

import logging
import re
import time
from pathlib import Path

import openpyxl

from .config import Config
from .models import ExtractedFinancials, RowLocation, WriteResult, Status

logger = logging.getLogger("tdnet")

# 保存リトライの待機秒数（フィボナッチ風）
RETRY_DELAYS = [3, 5, 8, 13, 21]


def _col_letter_to_index(col: str) -> int:
    """Excel列文字 → 1-indexed の列番号（A=1, B=2, ...）"""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _cell_value_str(ws, row: int, col_letter: str) -> str:
    """ワークシートのセル値を文字列で取得"""
    val = ws[f"{col_letter}{row}"].value
    if val is None:
        return ""
    return str(val).strip()


def _is_cell_empty(ws, row: int, col_letter: str) -> bool:
    """セルが空かどうか"""
    val = ws[f"{col_letter}{row}"].value
    return val is None or str(val).strip() == ""


def find_target_row(
    ws,
    code: str,
    fiscal_term: str,
    quarter: str,
    config: Config,
) -> RowLocation | WriteResult:
    """
    行特定ロジック（誤爆完全防止）

    Returns:
        RowLocation（成功時）or WriteResult（エラー時、status付き）
    """
    cols = config.columns

    # ============================================================
    # ① A列でコード検索 → startRow
    # ============================================================
    start_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = _cell_value_str(ws, row, cols.code)
        # 数値として比較（6750 == "6750"）
        if cell_val == str(code) or cell_val == code:
            start_row = row
            break

    if start_row is None:
        logger.warning(f"[Excel] コード '{code}' がA列に見つかりません")
        return WriteResult(
            status=Status.CODE_NOT_IN_SHEET,
            detail=f"コード '{code}' がA列({cols.code})に見つかりません",
        )

    logger.info(f"[Excel] コード '{code}' → startRow={start_row}")

    # ============================================================
    # ② M列でR年度探索（150行制限厳守）
    # 探索範囲: [startRow, startRow + max_scan_rows]
    # ============================================================
    term_row = None
    scan_end = min(start_row + config.max_scan_rows, ws.max_row + 1)

    for row in range(start_row, scan_end):
        cell_val = _cell_value_str(ws, row, cols.fiscal_term)
        if not cell_val:
            continue

        # 完全一致チェック（"R8/3" == "R8/3"）
        if cell_val == fiscal_term:
            term_row = row
            break

    if term_row is None:
        logger.warning(
            f"[Excel] 年度 '{fiscal_term}' がM列に見つかりません "
            f"(探索範囲: 行{start_row}〜{scan_end - 1})"
        )
        return WriteResult(
            status=Status.MISSING_TERM_WITHIN_150,
            detail=f"年度 '{fiscal_term}' が行{start_row}〜{scan_end - 1}の範囲で見つかりません",
        )

    logger.info(f"[Excel] 年度 '{fiscal_term}' → termRow={term_row}")

    # ============================================================
    # ③ N列で四半期探索（近傍限定）
    # 探索範囲: [termRow - q_search_up, termRow + q_search_down]
    # ============================================================
    target_row = None
    # 下限: startRow と (termRow - q_search_up) の大きい方で、別企業の行に入らない
    q_start = max(start_row, term_row - config.q_search_up)
    q_end = min(term_row + config.q_search_down, ws.max_row + 1)

    for row in range(q_start, q_end):
        cell_val = _cell_value_str(ws, row, cols.quarter)
        if cell_val == quarter:
            target_row = row
            break

    if target_row is None:
        logger.warning(
            f"[Excel] 四半期 '{quarter}' がN列に見つかりません "
            f"(探索範囲: 行{q_start}〜{q_end - 1})"
        )
        return WriteResult(
            status=Status.MISSING_QUARTER_NEAR_TERM,
            detail=f"四半期 '{quarter}' が行{q_start}〜{q_end - 1}の範囲で見つかりません",
        )

    logger.info(f"[Excel] 四半期 '{quarter}' → targetRow={target_row}")

    return RowLocation(start_row=start_row, term_row=term_row, target_row=target_row)


def _check_and_write_cell(
    ws, row: int, col_letter: str, new_value: int | None, field_name: str,
) -> tuple[str, str | None, int | None]:
    """
    セルの競合チェックと書き込み。

    Returns:
        (result, old_value_str, new_value)
        result: "written" | "skipped_same" | "skipped_null" | "conflict"
    """
    if new_value is None:
        return "skipped_null", None, None

    old_val = ws[f"{col_letter}{row}"].value
    old_str = str(old_val).strip() if old_val is not None else ""

    # 空セル → 書き込み
    if old_val is None or old_str == "":
        ws[f"{col_letter}{row}"] = new_value
        logger.info(f"[Excel] {field_name}: 行{row}{col_letter}列 に {new_value} を書き込み")
        return "written", "", new_value

    # 旧値と新値を比較
    try:
        old_num = int(float(str(old_val)))
    except (ValueError, TypeError):
        old_num = None

    if old_num == new_value:
        logger.info(f"[Excel] {field_name}: 行{row}{col_letter}列 は既に同値 ({new_value})、スキップ")
        return "skipped_same", old_str, new_value

    # 旧値 ≠ 新値 → 上書き禁止！
    logger.warning(
        f"[Excel] {field_name}: 行{row}{col_letter}列 で競合検知！"
        f" 旧値={old_str}, 新値={new_value} → 上書き禁止"
    )
    return "conflict", old_str, new_value


def write_to_excel(
    config: Config,
    code: str,
    financials: ExtractedFinancials,
) -> WriteResult:
    """
    Excelに決算数値を安全に書き込む。

    安全保証:
    - 150行制限・近傍限定で誤爆防止
    - 競合検知で上書き禁止
    - 保存リトライでロック対応
    """
    excel_path = config.excel_path
    cols = config.columns

    if not Path(excel_path).exists():
        return WriteResult(
            status=Status.FILE_LOCKED_OR_SAVE_FAILED,
            detail=f"Excelファイルが見つかりません: {excel_path}",
        )

    # ① ファイルを開く（最新状態を読み込み）
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        return WriteResult(
            status=Status.FILE_LOCKED_OR_SAVE_FAILED,
            detail=f"Excelファイルを開けません: {e}",
        )

    # シート取得
    if config.sheet_name not in wb.sheetnames:
        wb.close()
        return WriteResult(
            status=Status.CODE_NOT_IN_SHEET,
            detail=f"シート '{config.sheet_name}' が見つかりません",
        )

    ws = wb[config.sheet_name]

    # ② 行特定
    location = find_target_row(
        ws, code, financials.fiscal_year, financials.quarter, config,
    )

    if isinstance(location, WriteResult):
        # エラー（行特定失敗）
        wb.close()
        return location

    target_row = location.target_row

    # ③ 旧値取得 & 競合チェック & 書き込み
    old_values: dict[str, str] = {}
    new_values: dict[str, int | None] = {}
    any_conflict = False
    any_written = False

    # 売上 (O列)
    result_s, old_s, new_s = _check_and_write_cell(
        ws, target_row, cols.sales, financials.sales, "売上高",
    )
    old_values["sales"] = old_s or ""
    new_values["sales"] = new_s
    if result_s == "conflict":
        any_conflict = True
    elif result_s == "written":
        any_written = True

    # 粗利益 (P列) — NULLならスキップ
    result_g, old_g, new_g = _check_and_write_cell(
        ws, target_row, cols.gross_profit, financials.gross_profit, "粗利益",
    )
    old_values["gross_profit"] = old_g or ""
    new_values["gross_profit"] = new_g
    if result_g == "conflict":
        any_conflict = True
    elif result_g == "written":
        any_written = True

    # 営業利益 (S列)
    result_o, old_o, new_o = _check_and_write_cell(
        ws, target_row, cols.operating_profit, financials.operating_profit, "営業利益",
    )
    old_values["operating_profit"] = old_o or ""
    new_values["operating_profit"] = new_o
    if result_o == "conflict":
        any_conflict = True
    elif result_o == "written":
        any_written = True

    # 競合があれば全体を中止（書き込み済みセルも保存しない）
    if any_conflict:
        wb.close()
        return WriteResult(
            status=Status.CONFLICT_DETECTED,
            target_row=target_row,
            old_values=old_values,
            new_values=new_values,
            detail="既存値との競合を検知。上書き禁止により書き込みを中止。",
        )

    # 書き込みがなかった場合（全て同値 or 全てNULL）
    if not any_written:
        wb.close()
        return WriteResult(
            status=Status.SUCCESS,
            target_row=target_row,
            old_values=old_values,
            new_values=new_values,
            detail="全セル同値またはNULL。書き込み不要。",
        )

    # ④ 保存リトライ
    retry_delays = RETRY_DELAYS[:config.retry_count]
    for attempt, delay in enumerate(retry_delays):
        try:
            wb.save(excel_path)
            wb.close()
            logger.info(f"[Excel] 保存成功（行{target_row}）")
            return WriteResult(
                status=Status.SUCCESS,
                target_row=target_row,
                old_values=old_values,
                new_values=new_values,
                detail=f"行{target_row}に書き込み成功",
            )
        except PermissionError:
            logger.warning(
                f"[Excel] 保存失敗（ロック中）、リトライ {attempt + 1}/{len(retry_delays)} "
                f"（{delay}秒後）"
            )
            time.sleep(delay)
        except Exception as e:
            logger.error(f"[Excel] 保存失敗: {e}")
            wb.close()
            return WriteResult(
                status=Status.FILE_LOCKED_OR_SAVE_FAILED,
                target_row=target_row,
                old_values=old_values,
                new_values=new_values,
                detail=f"保存失敗: {e}",
            )

    # 全リトライ失敗
    wb.close()
    return WriteResult(
        status=Status.FILE_LOCKED_OR_SAVE_FAILED,
        target_row=target_row,
        old_values=old_values,
        new_values=new_values,
        detail=f"保存リトライ{len(retry_delays)}回全て失敗（ファイルロック）",
    )
