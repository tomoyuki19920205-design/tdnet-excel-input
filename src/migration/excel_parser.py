# ============================================================
# excel_parser.py — Excel業績シートパーサー（高速版）
# ============================================================
"""
1シート内に縦並びで1000社以上の業績データが入った
Excelファイルを上→下に走査し、企業ブロック単位でパースする。

仕様:
  - A列: 企業コード（ブロック開始マーカー）
  - B列: 無視
  - C〜L列: 補助メモ（文字列そのまま保存）— 企業コード行から読み取り
  - M列: 年度（1Q行にのみ存在、2Q〜4Qは引継ぎ）
  - N列: 四半期（1Q/2Q/3Q/4Q のみ有効）
  - O列: 売上, P列: 粗利, Q列: 粗利率, R列: 管理費, S列: 営業利益
  - Z列: 四半期メモ（改行保持）
  - AA列〜: セグメント（売上/利益ペア）
  - ヘッダー行: O=売上, P=粗利, Q=粗利率, R=管理費

パフォーマンス最適化:
  - ws.iter_rows(values_only=True) を1回だけ呼び、全行を単一パスで処理
  - read_onlyモードでのiter_rows再呼出し(XML先頭再パース)を完全排除
  - A〜max_col を一括読み取り、Z列・セグメントはQ行でのみ値を使用
  - 150行制限で企業ブロックを早期終了
"""
from __future__ import annotations

import calendar
import logging
import re
from typing import Any

import openpyxl

from .parse_models import (
    CompanyBlock,
    LogEntry,
    ParseResult,
    QuarterlyRecord,
    SegmentPair,
)

logger = logging.getLogger("migration")

# ------------------------------------------------------------------
# 定数
# ------------------------------------------------------------------
MAX_BLOCK_DISTANCE = 150   # 企業コード行からの最大探索距離

# 列番号（1-indexed）
COL_A = 1   # 企業コード
COL_C = 3   # 補助メモ開始
COL_L = 12  # 補助メモ終了
COL_M = 13  # 年度
COL_N = 14  # 四半期
COL_O = 15  # 売上
COL_P = 16  # 粗利
COL_Q = 17  # 粗利率
COL_R = 18  # 管理費
COL_S = 19  # 営業利益
COL_Z = 26  # 四半期メモ
COL_AA = 27 # セグメント開始

# タプルインデックス（0-indexed）
IDX_A = 0   # A列
IDX_C = 2   # C列
IDX_L = 11  # L列
IDX_M = 12  # M列
IDX_N = 13  # N列
IDX_O = 14  # O列
IDX_P = 15  # P列
IDX_Q = 16  # Q列
IDX_R = 17  # R列
IDX_S = 18  # S列
IDX_Z = 25  # Z列
IDX_AA = 26 # AA列（セグメント開始）

# 有効な四半期値
VALID_QUARTERS = {"1Q", "2Q", "3Q", "4Q"}

# 全角→半角 変換テーブル（数字＋英大文字＋英小文字）
_ZEN2HAN = str.maketrans(
    "０１２３４５６７８９"
    "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
    "0123456789"
    "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    "abcdefghijklmnopqrstuvwxyz",
)


# ------------------------------------------------------------------
# ユーティリティ（値ベース — wsアクセスなし）
# ------------------------------------------------------------------
def _val_str(val: Any) -> str | None:
    """値を文字列で取得。None/空 → None"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _val_num(val: Any) -> float | None:
    """値を数値で取得。非数値 → None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    # 文字列の場合: カンマ・全角除去して数値化を試みる
    s = str(val).strip()
    s = s.replace(",", "").replace("，", "")
    s = s.replace("　", "")
    s = s.translate(_ZEN2HAN)
    # マイナス記号の正規化
    if s.startswith("△") or s.startswith("▲"):
        s = "-" + s[1:]
    s = s.replace("－", "-").replace("‐", "-")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _detect_company_code(val: Any) -> str | None:
    """
    A列の値から企業コードを検出する。
    企業コードは4〜5桁の数字。
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # 全角→半角
    s = s.translate(_ZEN2HAN)
    # 4〜5桁の数字のみ
    m = re.match(r"^(\d{4,5})$", s)
    if m:
        return m.group(1)
    return None


def _normalize_quarter(val: Any) -> str | None:
    """
    N列の値を正規化して四半期文字列を返す。
    有効値: 1Q / 2Q / 3Q / 4Q のみ。

    対応パターン: 1Q, ２Ｑ, 3Ｑ, ４ｑ, 1, ２ など
    全角英数字・大文字小文字をすべて正規化。
    """
    if val is None:
        return None
    s = str(val).strip()
    # 全角→半角（数字＋英字）
    s = s.translate(_ZEN2HAN)
    s = s.upper()
    # "1Q" 〜 "4Q" フォーマット
    if s in VALID_QUARTERS:
        return s
    # 数字だけの場合 "1" → "1Q"
    if s in ("1", "2", "3", "4"):
        return f"{s}Q"
    return None


def _is_header_row(row_values: tuple) -> bool:
    """
    ヘッダー行を判定する（タプル版・緩和条件）。

    必須条件:
      O列に「売上」または「売上収益」を含む（部分一致）
    補助条件（以下のうち2つ以上一致でヘッダー確定）:
      P列に「粗利」を含む
      Q列に「粗利率」を含む
      R列に「管理費」を含む

    O列必須 + 補助2/3以上 → ヘッダー行と判定。
    O列必須 + 補助0〜1 → ヘッダーではない。
    """
    if len(row_values) < IDX_R + 1:
        return False

    # O列: 必須チェック（「売上」または「売上収益」を含む）
    o_val = row_values[IDX_O]
    if o_val is None:
        return False
    o_str = str(o_val).strip()
    if not o_str or ("売上" not in o_str):
        return False

    # P/Q/R列: 補助チェック（3つのうち2つ以上一致）
    aux_matches = 0
    aux_checks = [
        (IDX_P, "粗利"),
        (IDX_Q, "粗利率"),
        (IDX_R, "管理費"),
    ]
    for idx, keyword in aux_checks:
        val = row_values[idx]
        if val is not None:
            s = str(val).strip()
            if s and keyword in s:
                aux_matches += 1

    return aux_matches >= 2


def _parse_fiscal_year_end(raw_year: str | None) -> str | None:
    """
    M列の年度文字列を YYYY-MM-DD（末日）形式に正規化する。

    対応パターン（完全一致のみ、部分一致禁止）:
      - "R8/3" → "2026-03-31"
      - "2026/3" → "2026-03-31"
      - "2026年3月" → "2026-03-31"
      - "令和8年3月期" → "2026-03-31"
      - "R8/3期" → "2026-03-31"

    月は1〜12の範囲のみ許可。範囲外やノイズ文字列はNoneを返す（例外なし）。
    """
    if raw_year is None:
        return None
    s = str(raw_year).strip()
    if not s:
        return None

    # 全角→半角
    s = s.translate(_ZEN2HAN)
    # "期"を除去
    s = s.replace("期", "")

    year: int | None = None
    month: int | None = None

    # R表記: R8/3（完全一致）
    m = re.match(r"^R(\d+)/(\d{1,2})$", s, re.IGNORECASE)
    if m:
        year = 2018 + int(m.group(1))
        month = int(m.group(2))

    # 西暦スラッシュ: 2026/3（完全一致）
    if year is None:
        m = re.match(r"^(\d{4})/(\d{1,2})$", s)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))

    # 西暦漢字: 2026年3月（完全一致）
    if year is None:
        m = re.match(r"^(\d{4})年(\d{1,2})月$", s)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))

    # 令和漢字: 令和8年3月（完全一致）
    if year is None:
        m = re.match(r"^令和(\d+)年(\d{1,2})月$", s)
        if m:
            year = 2018 + int(m.group(1))
            month = int(m.group(2))

    # いずれにもマッチしない or 月が不正
    if year is None or month is None:
        return None
    if not (1 <= month <= 12):
        return None

    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-{last_day:02d}"



def _read_memo_from_tuple(row_values: tuple) -> dict[str, str | None]:
    """C〜L列の補助メモをタプルから読み取る。"""
    memo = {}
    col_names = ["c", "d", "e", "f", "g", "h", "i", "j", "k", "l"]
    for i, col_name in enumerate(col_names):
        idx = IDX_C + i  # 2, 3, ..., 11
        if idx < len(row_values):
            val = row_values[idx]
        else:
            val = None
        if val is None:
            memo[f"memo_{col_name}"] = None
        else:
            memo[f"memo_{col_name}"] = str(val)  # 改行保持、数値も文字列化
    return memo


def _parse_segments_from_tuples(
    data_values: tuple,
    header_values: tuple | None,
) -> list[SegmentPair]:
    """
    セグメントデータをタプルから読み取る。
    data_values / header_values は AA列以降のスライス済みタプル。
    """
    segments: list[SegmentPair] = []
    if data_values is None:
        return segments

    max_len = len(data_values)
    if max_len == 0:
        return segments

    header_len = len(header_values) if header_values else 0
    col = 0
    order = 0

    while col < max_len:
        # ヘッダー行からセグメント名を取得
        seg_name = None
        if header_values is not None and col < header_len:
            raw = header_values[col]
            if raw is not None:
                s = str(raw).strip()
                seg_name = s if s else None

        # 売上値
        seg_sales = _val_num(data_values[col])
        # 利益値（次列）
        seg_profit = None
        if col + 1 < max_len:
            seg_profit = _val_num(data_values[col + 1])

        # 売上も利益もNoneならスキップ
        if seg_sales is None and seg_profit is None:
            col += 2
            order += 1
            continue

        # セグメント名の決定
        if seg_name is None:
            seg_name = f"UNKNOWN_{order + 1}"

        segments.append(SegmentPair(
            segment_name=seg_name,
            segment_order=order,
            segment_sales=seg_sales,
            segment_profit=seg_profit,
        ))

        col += 2  # 売上/利益ペアなので2列ずつ進む
        order += 1

    return segments


# ------------------------------------------------------------------
# メインパーサー
# ------------------------------------------------------------------
def parse_excel(
    file_path: str,
    sheet_name: str = "PL",
) -> ParseResult:
    """
    Excelファイルの1シートを上→下に走査し、企業ブロック単位でパースする。

    高速化ポイント:
      - iter_rows(values_only=True) を1回だけ呼び全行を単一パスで処理
      - read_onlyモードでのiter_rows再呼出し(XML先頭再パース)を完全排除
      - Z列・セグメントはQ行でのみタプルから値を使用
      - 150行制限で企業ブロックを早期終了

    Returns:
        ParseResult — ブロック一覧 + ログ一覧
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        raise ValueError(f"シート '{sheet_name}' が見つかりません。"
                         f" 利用可能: {available}")
    ws = wb[sheet_name]

    result = ParseResult()
    max_row = ws.max_row or 0
    max_col = ws.max_column or COL_S
    result.total_rows_scanned = max_row

    # セグメント列が存在するか
    has_segments = max_col >= COL_AA

    # --- ブロック状態変数 ---
    current_code: str | None = None
    block_start_row: int = 0
    header_row_values: tuple | None = None   # ヘッダー行の全タプル
    current_fiscal_year: str | None = None
    current_memo: dict[str, str | None] = {}
    records: list[QuarterlyRecord] = []
    rows_since_code: int = 0
    block_has_header: bool = False

    def _flush_block() -> None:
        """現在のブロックを結果に追加する。"""
        nonlocal current_code, records, header_row_values
        nonlocal current_fiscal_year, current_memo, rows_since_code
        nonlocal block_has_header

        if current_code and records:
            block = CompanyBlock(
                company_code=current_code,
                row_start=block_start_row,
                row_end=block_start_row + rows_since_code,
                **current_memo,  # type: ignore
            )
            block.records = records
            result.blocks.append(block)
        elif current_code and not records:
            block = CompanyBlock(
                company_code=current_code,
                row_start=block_start_row,
                row_end=block_start_row + rows_since_code,
                **current_memo,  # type: ignore
            )
            block.records = []
            result.blocks.append(block)

        # リセット
        current_code = None
        header_row_values = None
        current_fiscal_year = None
        current_memo = {}
        records = []
        rows_since_code = 0
        block_has_header = False

    # ================================================================
    # 単一パス: A〜max_col を1回のiter_rowsで全行走査
    # ================================================================
    # read_onlyモードではiter_rowsを複数回呼ぶとXMLを毎回先頭から
    # 再パースするため、必ず1回のイテレーションで全て処理する。
    row_num = 0
    for row_values in ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=COL_A,
        max_col=max_col,
        values_only=True,
    ):
        row_num += 1

        # --- A列: 企業コード検出 ---
        code = _detect_company_code(row_values[IDX_A])

        if code is not None:
            # 前ブロックのフラッシュ
            if current_code is not None:
                _flush_block()

            current_code = code
            block_start_row = row_num
            rows_since_code = 0
            # C〜L列の補助メモを読み取り（企業コード行から）
            current_memo = _read_memo_from_tuple(row_values)

        if current_code is None:
            continue

        rows_since_code += 1

        # --- 150行超過チェック ---
        if rows_since_code > MAX_BLOCK_DISTANCE:
            result.logs.append(LogEntry(
                log_level="SKIP",
                log_type="SKIP_DISTANCE",
                message=f"企業コード {current_code}: {MAX_BLOCK_DISTANCE}行超過",
                row_start=block_start_row,
                row_end=row_num,
                company_code=current_code,
            ))
            _flush_block()
            continue

        # --- ヘッダー検知（O〜R列の4値のみチェック） ---
        if not block_has_header:
            if _is_header_row(row_values):
                block_has_header = True
                # ヘッダー行の全タプルを保持（セグメント名取得用）
                header_row_values = row_values
            continue  # ヘッダー未検知 or ヘッダー行自体はデータでない

        # --- N列: 四半期判定（早期continue） ---
        q = _normalize_quarter(row_values[IDX_N])
        if q is None:
            continue

        # --- 年度決定 ---
        if q == "1Q":
            raw_year = _val_str(row_values[IDX_M])
            fy = _parse_fiscal_year_end(raw_year)
            if fy is None and raw_year is not None:
                result.logs.append(LogEntry(
                    log_level="WARN",
                    log_type="WARN_YEAR_PARSE",
                    message=f"年度解析不可(スキップ): '{raw_year}' (row={row_num})",
                    row_start=row_num,
                    row_end=row_num,
                    company_code=current_code,
                    quarter=q,
                ))
                continue
            if fy is not None:
                current_fiscal_year = fy

        if current_fiscal_year is None:
            result.logs.append(LogEntry(
                log_level="SKIP",
                log_type="SKIP_NO_QUARTER",
                message=f"年度不明のためスキップ (row={row_num}, Q={q})",
                row_start=row_num,
                row_end=row_num,
                company_code=current_code,
                quarter=q,
            ))
            continue

        # --- PL数値読み取り（row_values の O〜S 列） ---
        sales = _val_num(row_values[IDX_O])
        gross_profit = _val_num(row_values[IDX_P])
        gross_margin = _val_num(row_values[IDX_Q])
        sga = _val_num(row_values[IDX_R])
        operating_profit = _val_num(row_values[IDX_S])

        # --- Z列メモ（同じ row_values から取得、Q行でのみ到達） ---
        note = None
        if IDX_Z < len(row_values):
            note = _val_str(row_values[IDX_Z])

        rec = QuarterlyRecord(
            company_code=current_code,
            fiscal_year_end=current_fiscal_year,
            quarter=q,
            row_number=row_num,
            sales=sales,
            gross_profit=gross_profit,
            gross_margin=gross_margin,
            sga=sga,
            operating_profit=operating_profit,
            note=note,
        )

        # --- セグメント読み取り（同じ row_values のAA列以降をスライス） ---
        if has_segments and header_row_values is not None:
            seg_data = row_values[IDX_AA:] if IDX_AA < len(row_values) else ()
            seg_header = header_row_values[IDX_AA:] if IDX_AA < len(header_row_values) else ()
            if seg_data:
                rec.segments = _parse_segments_from_tuples(
                    data_values=seg_data,
                    header_values=seg_header,
                )

        records.append(rec)

    # 最後のブロックをフラッシュ
    if current_code is not None:
        _flush_block()

    wb.close()
    return result
