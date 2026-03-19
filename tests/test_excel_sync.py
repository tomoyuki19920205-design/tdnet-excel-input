#!/usr/bin/env python3
"""tests/test_excel_sync.py — Excel同期ツールのテスト"""
from __future__ import annotations

import os
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

# --- テスト対象 ---
import sys
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from tools.excel_sync import (
    push_to_excel,
    pull_from_excel,
    _parse_fy_label,
    _parse_quarter_label,
    _upsert_memo,
    _jpy_to_display,
    _get_company_id,
    _get_period_id,
)


# ============================================================
# フィクスチャ
# ============================================================

@pytest.fixture
def db_and_excel(tmp_path):
    """
    テスト用DBとExcelファイルを作成する。
    DB: companies(5461), periods(2025-12-31 Q2), facts(NET_SALES, OP_INCOME)
    Excel: A/M/N列にキーデータ、O〜は空
    """
    db_path = str(tmp_path / "test.db")

    # --- DB作成 ---
    schema_path = os.path.join(_PROJECT_ROOT, "schema.sql")
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    with open(schema_path, "r", encoding="utf-8") as f:
        conn.executescript(f.read())

    # company
    conn.execute(
        "INSERT INTO companies (company_id, ticker_code, name_ja) VALUES (1, '5461', 'テスト株式会社')"
    )
    conn.execute(
        "INSERT INTO companies (company_id, ticker_code, name_ja) VALUES (2, '2288', 'テスト2')"
    )

    # periods
    conn.execute(
        "INSERT INTO periods (period_id, company_id, fiscal_year_end, fiscal_year, quarter) "
        "VALUES (1, 1, '2026-03-31', 2026, 2)"
    )
    conn.execute(
        "INSERT INTO periods (period_id, company_id, fiscal_year_end, fiscal_year, quarter) "
        "VALUES (2, 2, '2026-03-31', 2026, 1)"
    )

    # disclosure (facts用のFK)
    conn.execute(
        "INSERT INTO disclosures (disclosure_id, company_id, source, disclosed_at, title, doc_type) "
        "VALUES (1, 1, 'TDNET', '2026-02-26 12:00:00', 'テスト決算短信', 'TANSHIN')"
    )
    conn.execute(
        "INSERT INTO disclosures (disclosure_id, company_id, source, disclosed_at, title, doc_type) "
        "VALUES (2, 2, 'TDNET', '2026-02-26 12:00:00', 'テスト2決算短信', 'TANSHIN')"
    )

    # facts (円整数)
    facts = [
        (1, 1, 1, "CONSOLIDATED", "NET_SALES",   36809000000, "JPY", "IXBRL"),
        (1, 1, 1, "CONSOLIDATED", "GROSS_PROFIT", 5000000000, "JPY", "IXBRL"),
        (1, 1, 1, "CONSOLIDATED", "OP_INCOME",     735000000, "JPY", "IXBRL"),
        (2, 2, 2, "CONSOLIDATED", "NET_SALES",  183643000000, "JPY", "IXBRL"),
        (2, 2, 2, "CONSOLIDATED", "OP_INCOME",    6807000000, "JPY", "IXBRL"),
    ]
    for cid, pid, did, scope, metric, val, unit, quality in facts:
        conn.execute(
            "INSERT INTO facts (company_id, period_id, disclosure_id, scope, metric, value, unit, quality) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (cid, pid, did, scope, metric, val, unit, quality),
        )

    # segment_facts
    conn.execute(
        "INSERT INTO segment_facts (company_id, period_id, segment_name, segment_order, sales, profit) "
        "VALUES (1, 1, '鉄鋼', 0, 20000000000, 500000000)"
    )
    conn.execute(
        "INSERT INTO segment_facts (company_id, period_id, segment_name, segment_order, sales, profit) "
        "VALUES (1, 1, '商社', 1, 16809000000, 235000000)"
    )

    conn.commit()
    conn.close()

    # --- Excel作成 ---
    excel_path = str(tmp_path / "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL"

    # ヘッダー (row 1)
    headers = {
        "A": "企業コード", "M": "会計年", "N": "四半期",
        "O": "累計売上", "P": "累計粗利", "Q": "粗利率",
        "R": "販管費", "S": "累計営業利益", "T": "営利率",
        "U": "単体売上", "V": "単体粗利", "W": "粗利率(単)", "X": "販管費(単)",
        "Y": "単体営益", "Z": "メモ",
        "AA": "セグA売上", "AB": "セグA利益", "AC": "セグB売上", "AD": "セグB利益",
    }
    for col_letter, header in headers.items():
        col_idx = _col_to_idx(col_letter)
        ws.cell(row=1, column=col_idx, value=header)

    # Data rows
    # row 2: 5461, R8/3, 2Q
    ws.cell(row=2, column=1, value="5461")
    ws.cell(row=2, column=13, value="R8/3")
    ws.cell(row=2, column=14, value="2Q")

    # row 3: 2288, R8/3, 1Q
    ws.cell(row=3, column=1, value="2288")
    ws.cell(row=3, column=13, value="R8/3")
    ws.cell(row=3, column=14, value="1Q")

    # row 4: 不明企業（DB照合不可）
    ws.cell(row=4, column=1, value="9999")
    ws.cell(row=4, column=13, value="R8/3")
    ws.cell(row=4, column=14, value="1Q")

    wb.save(excel_path)

    return db_path, excel_path


def _col_to_idx(col_letter: str) -> int:
    """列文字 → 1-indexed"""
    from tools.excel_sync import _COL
    return _COL.get(col_letter, 0) + 1


# ============================================================
# ユーティリティ関数テスト
# ============================================================

class TestParseHelpers:
    def test_parse_fy_label_r8_3(self):
        result = _parse_fy_label("R8/3")
        assert result == ("2026-03-31", 2026)

    def test_parse_fy_label_r7_12(self):
        result = _parse_fy_label("R7/12")
        assert result == ("2025-12-31", 2025)

    def test_parse_fy_label_invalid(self):
        assert _parse_fy_label("2025") is None
        assert _parse_fy_label("") is None

    def test_parse_quarter_label(self):
        assert _parse_quarter_label("1Q") == 1
        assert _parse_quarter_label("2Q") == 2
        assert _parse_quarter_label("4Q") == 4
        assert _parse_quarter_label("5Q") is None
        assert _parse_quarter_label("") is None

    def test_jpy_to_display_million(self):
        assert _jpy_to_display(1_000_000, "百万円") == 1.0
        assert _jpy_to_display(36_809_000_000, "百万円") == 36809.0
        assert _jpy_to_display(None, "百万円") is None

    def test_jpy_to_display_yen(self):
        assert _jpy_to_display(1000, "円") == 1000.0


# ============================================================
# push テスト (DB→Excel)
# ============================================================

class TestPush:
    def test_basic_push(self, db_and_excel):
        """push後にExcelのO列に売上値が入る"""
        db_path, excel_path = db_and_excel
        result = push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        assert result["rows_updated"] >= 1

        # Excelを読み返す
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]

        # row 2: 5461, R8/3, 2Q → NET_SALES 36809百万円
        o_val = ws.cell(row=2, column=_col_to_idx("O")).value
        assert o_val == 36809.0

        # P列: GROSS_PROFIT 5000百万円
        p_val = ws.cell(row=2, column=_col_to_idx("P")).value
        assert p_val == 5000.0

        # S列: OP_INCOME 735百万円
        s_val = ws.cell(row=2, column=_col_to_idx("S")).value
        assert s_val == 735.0

    def test_formula_columns(self, db_and_excel):
        """push後にQ,R,T列にExcel数式が設定される"""
        db_path, excel_path = db_and_excel
        push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]

        # Q列: =IF(O2=0,"",P2/O2)
        q_val = ws.cell(row=2, column=_col_to_idx("Q")).value
        assert "P2" in str(q_val) and "O2" in str(q_val)

        # R列: =P2-S2
        r_val = ws.cell(row=2, column=_col_to_idx("R")).value
        assert "P2" in str(r_val) and "S2" in str(r_val)

        # T列: =IF(O2=0,"",S2/O2)
        t_val = ws.cell(row=2, column=_col_to_idx("T")).value
        assert "S2" in str(t_val) and "O2" in str(t_val)

    def test_segment_columns(self, db_and_excel):
        """push後にAA/AB列にセグメントデータが書き込まれる"""
        db_path, excel_path = db_and_excel
        push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]

        # row 2: 5461 のセグメント
        # AA列(27): 鉄鋼 売上 20000百万円
        aa_val = ws.cell(row=2, column=27).value
        assert aa_val == 20000.0

        # AB列(28): 鉄鋼 利益 500百万円
        ab_val = ws.cell(row=2, column=28).value
        assert ab_val == 500.0

        # AC列(29): 商社 売上 16809百万円
        ac_val = ws.cell(row=2, column=29).value
        assert ac_val == 16809.0

        # AD列(30): 商社 利益 235百万円
        ad_val = ws.cell(row=2, column=30).value
        assert ad_val == 235.0

    def test_multiple_companies(self, db_and_excel):
        """複数企業の行がそれぞれ正しく更新される"""
        db_path, excel_path = db_and_excel
        result = push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        assert result["rows_updated"] >= 2

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]

        # row 3: 2288
        o_val = ws.cell(row=3, column=_col_to_idx("O")).value
        assert o_val == 183643.0

    def test_unknown_company_skipped(self, db_and_excel):
        """DBにない企業コードの行はスキップされる"""
        db_path, excel_path = db_and_excel
        push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]

        # row 4: 9999 は更新されない
        o_val = ws.cell(row=4, column=_col_to_idx("O")).value
        assert o_val is None

    def test_push_idempotent(self, db_and_excel):
        """同じpushを2回実行しても値が壊れない"""
        db_path, excel_path = db_and_excel
        push_to_excel(db_path, excel_path, "PL", "百万円", 10)
        push_to_excel(db_path, excel_path, "PL", "百万円", 10)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        assert ws.cell(row=2, column=_col_to_idx("O")).value == 36809.0


# ============================================================
# pull テスト (Excel→DB, Z列メモ)
# ============================================================

class TestPull:
    def test_memo_saved_to_db(self, db_and_excel):
        """Z列メモがDBに保存される"""
        db_path, excel_path = db_and_excel

        # Excelにメモを書き込む
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        ws.cell(row=2, column=_col_to_idx("Z"), value="好決算！")
        wb.save(excel_path)

        result = pull_from_excel(db_path, excel_path, "PL", 10)
        assert result["memos_saved"] == 1

        # DBを確認
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            "SELECT memo_text FROM quarterly_memos WHERE company_id = 1 AND period_id = 1"
        )
        row = cur.fetchone()
        assert row is not None
        assert row[0] == "好決算！"
        conn.close()

    def test_memo_updated(self, db_and_excel):
        """メモの2回目書き込みで更新される"""
        db_path, excel_path = db_and_excel

        # 1回目
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        ws.cell(row=2, column=_col_to_idx("Z"), value="初回メモ")
        wb.save(excel_path)
        pull_from_excel(db_path, excel_path, "PL", 10)

        # 2回目
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        ws.cell(row=2, column=_col_to_idx("Z"), value="更新メモ")
        wb.save(excel_path)
        result = pull_from_excel(db_path, excel_path, "PL", 10)
        assert result["memos_saved"] == 1

        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            "SELECT memo_text FROM quarterly_memos WHERE company_id = 1 AND period_id = 1"
        )
        assert cur.fetchone()[0] == "更新メモ"
        conn.close()

    def test_memo_unchanged_not_saved(self, db_and_excel):
        """同じメモで2回pullしたらno_change"""
        db_path, excel_path = db_and_excel

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        ws.cell(row=2, column=_col_to_idx("Z"), value="同じメモ")
        wb.save(excel_path)

        pull_from_excel(db_path, excel_path, "PL", 10)
        result = pull_from_excel(db_path, excel_path, "PL", 10)
        assert result["memos_unchanged"] == 1
        assert result["memos_saved"] == 0

    def test_memo_empty_skipped(self, db_and_excel):
        """空メモの行はDB保存しない"""
        db_path, excel_path = db_and_excel
        result = pull_from_excel(db_path, excel_path, "PL", 10)
        assert result["memos_saved"] == 0


# ============================================================
# DB操作テスト
# ============================================================

class TestDBOperations:
    def test_quarterly_memos_upsert(self, db_and_excel):
        """_upsert_memoのinsert/update/no_change"""
        db_path, _ = db_and_excel
        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA foreign_keys = ON")

        assert _upsert_memo(conn, 1, 1, "初回") == "inserted"
        conn.commit()

        assert _upsert_memo(conn, 1, 1, "初回") == "no_change"

        assert _upsert_memo(conn, 1, 1, "更新") == "updated"
        conn.commit()

        cur = conn.execute("SELECT memo_text FROM quarterly_memos WHERE company_id=1 AND period_id=1")
        assert cur.fetchone()[0] == "更新"

        conn.close()

    def test_schema_has_new_tables(self, db_and_excel):
        """schema.sqlにquarterly_memosとsegment_factsテーブルが存在する"""
        db_path, _ = db_and_excel
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        assert "quarterly_memos" in tables
        assert "segment_facts" in tables
        conn.close()
