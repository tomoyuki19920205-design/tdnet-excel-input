# ============================================================
# test_excel_writer.py — Excel書き込みのテスト（行特定・競合検知）
# ============================================================
import sys
import os
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

from src.config import Config, ColumnMapping
from src.models import ExtractedFinancials, RowLocation, WriteResult, Status
from src.excel_writer import find_target_row, write_to_excel


def _make_config(excel_path: str, **kwargs) -> Config:
    """テスト用Config"""
    return Config(
        excel_path=excel_path,
        sheet_name=kwargs.get("sheet_name", "PL"),
        max_scan_rows=kwargs.get("max_scan_rows", 150),
        q_search_up=kwargs.get("q_search_up", 20),
        q_search_down=kwargs.get("q_search_down", 40),
        retry_count=kwargs.get("retry_count", 1),
        columns=ColumnMapping(),
    )


def _create_test_excel(path: str, rows: list[dict]):
    """テスト用Excelファイルを作成する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL"

    for row_data in rows:
        row_num = row_data["row"]
        if "A" in row_data:
            ws[f"A{row_num}"] = row_data["A"]
        if "M" in row_data:
            ws[f"M{row_num}"] = row_data["M"]
        if "N" in row_data:
            ws[f"N{row_num}"] = row_data["N"]
        if "O" in row_data:
            ws[f"O{row_num}"] = row_data["O"]
        if "P" in row_data:
            ws[f"P{row_num}"] = row_data["P"]
        if "S" in row_data:
            ws[f"S{row_num}"] = row_data["S"]

    wb.save(path)
    wb.close()


class TestFindTargetRow:
    """find_target_row のテスト"""

    def test_success(self, tmp_path):
        """正常ケース: コード→年度→四半期の3段階で行特定"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "6750", "M": "R8/3", "N": "1Q"},
            {"row": 2, "A": "", "M": "R8/3", "N": "2Q"},
            {"row": 3, "A": "", "M": "R8/3", "N": "3Q"},
            {"row": 4, "A": "", "M": "R8/3", "N": "4Q"},
        ])

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path)

        result = find_target_row(ws, "6750", "R8/3", "2Q", config)
        assert isinstance(result, RowLocation)
        assert result.start_row == 1
        assert result.term_row == 1  # M列R8/3は行1から一致
        assert result.target_row == 2  # N列2Qは行2
        wb.close()

    def test_code_not_found(self, tmp_path):
        """コードがA列に無い場合"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "7203", "M": "R8/3", "N": "1Q"},
        ])

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path)

        result = find_target_row(ws, "6750", "R8/3", "1Q", config)
        assert isinstance(result, WriteResult)
        assert result.status == Status.CODE_NOT_IN_SHEET
        wb.close()

    def test_term_not_found_within_150(self, tmp_path):
        """年度がM列の150行範囲内に無い場合"""
        excel_path = str(tmp_path / "test.xlsx")
        rows = [{"row": 1, "A": "6750"}]
        # 行160にR8/3を設置（150行超 → 見つからないはず）
        rows.append({"row": 160, "M": "R8/3", "N": "1Q"})
        _create_test_excel(excel_path, rows)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path, max_scan_rows=150)

        result = find_target_row(ws, "6750", "R8/3", "1Q", config)
        assert isinstance(result, WriteResult)
        assert result.status == Status.MISSING_TERM_WITHIN_150
        wb.close()

    def test_term_found_within_150(self, tmp_path):
        """年度がM列の150行範囲内にある場合"""
        excel_path = str(tmp_path / "test.xlsx")
        rows = [{"row": 1, "A": "6750"}]
        rows.append({"row": 100, "M": "R8/3", "N": "2Q"})
        _create_test_excel(excel_path, rows)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path, max_scan_rows=150)

        result = find_target_row(ws, "6750", "R8/3", "2Q", config)
        assert isinstance(result, RowLocation)
        assert result.term_row == 100
        wb.close()

    def test_quarter_not_found_near_term(self, tmp_path):
        """四半期がN列の近傍に無い場合"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "6750", "M": "R8/3", "N": "1Q"},
        ])

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path)

        result = find_target_row(ws, "6750", "R8/3", "3Q", config)
        assert isinstance(result, WriteResult)
        assert result.status == Status.MISSING_QUARTER_NEAR_TERM
        wb.close()

    def test_multiple_companies(self, tmp_path):
        """複数企業が混在するシート"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "7203", "M": "R8/3", "N": "1Q"},
            {"row": 2, "A": "", "M": "R8/3", "N": "2Q"},
            {"row": 10, "A": "6750", "M": "R8/3", "N": "1Q"},
            {"row": 11, "A": "", "M": "R8/3", "N": "2Q"},
            {"row": 12, "A": "", "M": "R8/3", "N": "3Q"},
        ])

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        config = _make_config(excel_path)

        result = find_target_row(ws, "6750", "R8/3", "2Q", config)
        assert isinstance(result, RowLocation)
        assert result.start_row == 10
        assert result.target_row == 11
        wb.close()


class TestWriteToExcel:
    """write_to_excel のテスト"""

    def test_write_empty_cells(self, tmp_path):
        """空セルへの書き込み"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "6750", "M": "R8/3", "N": "2Q"},
        ])

        config = _make_config(excel_path)
        financials = ExtractedFinancials(
            sales=12345,
            gross_profit=None,  # 粗利はNULL → スキップ
            operating_profit=1234,
            fiscal_year="R8/3",
            quarter="2Q",
        )

        result = write_to_excel(config, "6750", financials)
        assert result.status == Status.SUCCESS

        # 書き込み結果を確認
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        assert ws["O1"].value == 12345
        assert ws["P1"].value is None  # 粗利はNULLスキップ
        assert ws["S1"].value == 1234
        wb.close()

    def test_conflict_detected(self, tmp_path):
        """既存値と異なる場合は上書き禁止"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "6750", "M": "R8/3", "N": "2Q", "O": 99999},
        ])

        config = _make_config(excel_path)
        financials = ExtractedFinancials(
            sales=12345,  # 99999 ≠ 12345 → conflict!
            operating_profit=1234,
            fiscal_year="R8/3",
            quarter="2Q",
        )

        result = write_to_excel(config, "6750", financials)
        assert result.status == Status.CONFLICT_DETECTED

        # 元の値が保持されていることを確認
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PL"]
        assert ws["O1"].value == 99999  # 上書きされていない！
        wb.close()

    def test_same_value_skip(self, tmp_path):
        """既存値と同じ場合はスキップ（success）"""
        excel_path = str(tmp_path / "test.xlsx")
        _create_test_excel(excel_path, [
            {"row": 1, "A": "6750", "M": "R8/3", "N": "2Q", "O": 12345, "S": 1234},
        ])

        config = _make_config(excel_path)
        financials = ExtractedFinancials(
            sales=12345,
            operating_profit=1234,
            fiscal_year="R8/3",
            quarter="2Q",
        )

        result = write_to_excel(config, "6750", financials)
        assert result.status == Status.SUCCESS

    def test_sheet_not_found(self, tmp_path):
        """シートが見つからない場合"""
        excel_path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(excel_path)
        wb.close()

        config = _make_config(excel_path, sheet_name="PL")
        financials = ExtractedFinancials(fiscal_year="R8/3", quarter="2Q")

        result = write_to_excel(config, "6750", financials)
        assert result.status == Status.CODE_NOT_IN_SHEET

    def test_150_row_limit_enforced(self, tmp_path):
        """150行制限が厳守されること"""
        excel_path = str(tmp_path / "test.xlsx")
        # A列にコードを行1に、M列にR8/3を行200に配置
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PL"
        ws["A1"] = "6750"
        ws["M200"] = "R8/3"
        ws["N200"] = "2Q"
        wb.save(excel_path)
        wb.close()

        config = _make_config(excel_path, max_scan_rows=150)
        financials = ExtractedFinancials(
            sales=12345,
            fiscal_year="R8/3",
            quarter="2Q",
        )

        result = write_to_excel(config, "6750", financials)
        # 行200は startRow(1) + 150 = 151 の範囲外 → 見つからない
        assert result.status == Status.MISSING_TERM_WITHIN_150
