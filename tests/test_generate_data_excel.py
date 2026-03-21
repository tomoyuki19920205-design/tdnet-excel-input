"""
test_generate_data_excel.py — data.xlsx 生成ツールのユニットテスト

現行 tools/generate_data_excel.py の API に追随:
  - _fye_to_iso (旧 _fye_to_label)
  - _quarter_label
  - _fetch_all
  - HEADERS (10列: ticker, period, quarter, sales, gross_profit,
             operating_profit, source, updated_at, recency_key, lookup_key)
  - SHEET_NAME
  - UserError
  - generate(output_path, supabase_url, supabase_key, sqlite_db)
"""
import os
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

# テスト対象
import sys
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from tools.generate_data_excel import (
    generate,
    _fye_to_iso,
    _quarter_label,
    _fetch_all,
    HEADERS,
    SHEET_NAME,
    UserError,
)


# ============================================================
# ヘルパー: モック応答生成
# ============================================================
def _mock_financials(n=3):
    """Supabase financials テーブル風のレコードを生成"""
    return [
        {
            "ticker": f"{i + 1:04d}",
            "period": "2026-03-31",
            "quarter": "2Q",
            "sales": (i + 1) * 1_000_000,
            "gross_profit": (i + 1) * 500_000,
            "operating_profit": (i + 1) * 200_000,
            "source": "tdnet",
            "updated_at": "2026-02-28T10:00:00+09:00",
        }
        for i in range(n)
    ]


def _mock_fetch_all_response(data_list):
    """_fetch_all のモック: 呼ばれたらdata_listを返す"""
    return data_list


# ============================================================
# テスト
# ============================================================

class TestFyeToIso:
    """旧 _fye_to_label → 現行 _fye_to_iso のテスト"""
    def test_normal(self):
        assert _fye_to_iso("2025/3") == "2025-03-31"

    def test_december(self):
        assert _fye_to_iso("2025/12") == "2025-12-31"

    def test_february(self):
        assert _fye_to_iso("2025/2") == "2025-02-28"

    def test_invalid(self):
        assert _fye_to_iso("invalid") == "invalid"


class TestQuarterLabel:
    def test_1q(self):
        assert _quarter_label("1") == "1Q"

    def test_4q(self):
        assert _quarter_label("4") == "4Q"

    def test_already_q(self):
        assert _quarter_label("2Q") == "2Q"


class TestGenerate:
    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_normal_generation(self, mock_ef, mock_seg, mock_fetch):
        """正常にdata.xlsxが生成される"""
        financials = _mock_financials(3)
        mock_fetch.return_value = financials

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            result = generate(
                output_path=out,
                supabase_url="https://test.supabase.co",
                supabase_key="test-key",
                sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
            )

            assert result["rows"] == 3
            assert result["errors"] == 0
            assert os.path.exists(out)

            # Excel内容確認
            wb = openpyxl.load_workbook(out)
            ws = wb[SHEET_NAME]

            # A1: 生成日時
            assert ws.cell(row=1, column=1).value.startswith("generated_at:")

            # ヘッダー確認 (2行目)
            for col, h in enumerate(HEADERS, 1):
                assert ws.cell(row=2, column=col).value == h

            # データ行確認 (3行目から)
            assert ws.cell(row=3, column=1).value is not None  # ticker
            assert ws.cell(row=3, column=4).value is not None  # sales

    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_empty_data_raises(self, mock_ef, mock_seg, mock_fetch):
        """空データではUserErrorが発生する（現行仕様）"""
        mock_fetch.return_value = []

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            with pytest.raises(UserError, match="データがありません"):
                generate(
                    output_path=out,
                    supabase_url="https://test.supabase.co",
                    supabase_key="test-key",
                    sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
                )

    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_ten_columns(self, mock_ef, mock_seg, mock_fetch):
        """10列が正しい順序で出力される"""
        mock_fetch.return_value = _mock_financials(1)

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            generate(
                output_path=out,
                supabase_url="https://test.supabase.co",
                supabase_key="test-key",
                sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
            )

            wb = openpyxl.load_workbook(out)
            ws = wb[SHEET_NAME]

            expected = [
                "ticker", "period", "quarter",
                "sales", "gross_profit", "operating_profit",
                "source", "updated_at", "recency_key", "lookup_key",
            ]
            for col, name in enumerate(expected, 1):
                assert ws.cell(row=2, column=col).value == name

    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_atomic_save(self, mock_ef, mock_seg, mock_fetch):
        """原子的保存：tempファイルが残らない"""
        mock_fetch.return_value = _mock_financials(1)

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            generate(
                output_path=out,
                supabase_url="https://test.supabase.co",
                supabase_key="test-key",
                sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
            )

            # 本体は存在
            assert os.path.exists(out)
            # tmpファイルは残っていない
            tmp_files = [f for f in os.listdir(tmpdir) if f.endswith(".tmp")]
            assert len(tmp_files) == 0

    @patch("tools.generate_data_excel._load_dotenv", return_value=False)
    @patch.dict(os.environ, {}, clear=True)
    def test_missing_credentials(self, mock_dotenv):
        """接続情報がない場合はUserErrorが出る"""
        with pytest.raises(UserError, match="接続情報が未設定"):
            generate(
                output_path="/tmp/test.xlsx",
                supabase_url="",
                supabase_key="",
            )

    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_recency_and_lookup_keys(self, mock_ef, mock_seg, mock_fetch):
        """recency_key と lookup_key が正しく生成される"""
        mock_fetch.return_value = [
            {
                "ticker": "7203",
                "period": "2025-03-31",
                "quarter": "4Q",
                "sales": 1000,
                "gross_profit": 500,
                "operating_profit": 200,
                "source": "jquants",
                "updated_at": "2025-05-01T00:00:00+09:00",
            }
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            generate(
                output_path=out,
                supabase_url="https://test.supabase.co",
                supabase_key="test-key",
                sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
            )

            wb = openpyxl.load_workbook(out)
            ws = wb[SHEET_NAME]

            # Row 3: data row
            assert ws.cell(row=3, column=9).value == "202504"   # recency_key
            assert ws.cell(row=3, column=10).value == "7203|2025-03-31|4Q"  # lookup_key

    @patch("tools.generate_data_excel._fetch_all")
    @patch("tools.generate_data_excel._read_segment_data", return_value=[])
    @patch("tools.generate_data_excel.write_extracted_facts_sheets",
           return_value={"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0})
    def test_result_includes_ticker_count(self, mock_ef, mock_seg, mock_fetch):
        """戻り値にtickers数が含まれる"""
        mock_fetch.return_value = _mock_financials(5)

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_data.xlsx")
            result = generate(
                output_path=out,
                supabase_url="https://test.supabase.co",
                supabase_key="test-key",
                sqlite_db=os.path.join(tmpdir, "nonexistent.db"),
            )

            assert result["tickers"] == 5
            assert result["rows"] == 5
