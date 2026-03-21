#!/usr/bin/env python3
"""
tests/test_generate_data_excel_extracted_facts.py
extracted_facts → data.xlsx 3シート (FORECAST_REVISION / MONTHLY_DATA / KPI_DATA) のテスト
"""
import os
import sqlite3
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.extraction.ir_doc_schema import ensure_tables, insert_document, insert_facts
from src.extraction.extracted_facts_sheets import (
    build_forecast_revision_rows,
    build_monthly_data_rows,
    build_kpi_data_rows,
    write_extracted_facts_sheets,
    write_sheet,
    FORECAST_SHEET, FORECAST_COLUMNS,
    MONTHLY_SHEET, MONTHLY_COLUMNS,
    KPI_SHEET, KPI_COLUMNS,
)


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
def db():
    """テスト用インメモリDB

    persist_policy 依存:
      insert_facts() は should_persist_intermediates() ガード付き。
      テスト環境ではデフォルト OFF のため、明示的に ON にする必要がある。
    """
    from src.persist_policy import init_persist_policy, reset_persist_policy
    init_persist_policy(cli_flag=True)
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    ensure_tables(conn)
    yield conn
    conn.close()
    reset_persist_policy()


def _add_doc(conn, ticker, pubdate, title, doc_type, file_type="html"):
    """テスト用 document 追加ヘルパー"""
    return insert_document(
        conn, ticker=ticker, pubdate=pubdate,
        title=title, doc_type=doc_type, file_type=file_type,
        url=f"https://test/{ticker}_{doc_type}_{pubdate}.html",
    )


def _add_fact(conn, document_id, ticker, metric_name, metric_value,
              period="2027-03-31", quarter="4Q", unit="円",
              source_type="html", confidence="medium",
              segment_name="", raw_label="", normalized_label="",
              table_title="", page_no=None):
    """テスト用 fact 追加ヘルパー"""
    insert_facts(conn, [{
        "document_id": document_id,
        "ticker": ticker,
        "period": period,
        "quarter": quarter,
        "metric_name": metric_name,
        "metric_value": metric_value,
        "unit": unit,
        "source_type": source_type,
        "confidence": confidence,
        "segment_name": segment_name,
        "raw_label": raw_label,
        "normalized_label": normalized_label,
        "table_title": table_title,
        "page_no": page_no,
    }])


# ============================================================
# FORECAST_REVISION Tests
# ============================================================

class TestForecastRevision:

    def test_before_after_pairing(self, db):
        """before/after の raw_label ペアリングが機能する"""
        doc_id = _add_doc(db, "4062", "2027-01-15", "業績予想の修正", "forecast_revision")
        _add_fact(db, doc_id, "4062", "sales", 100_000_000_000,
                  raw_label="前回予想:売上高")
        _add_fact(db, doc_id, "4062", "sales", 120_000_000_000,
                  raw_label="今回予想:売上高")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["before_value"] == 100_000_000_000
        assert r["after_value"] == 120_000_000_000
        assert r["ticker"] == "4062"
        assert r["doc_type"] == "forecast_revision"

    def test_delta_calculation(self, db):
        """delta_value と delta_pct が正しく計算される"""
        doc_id = _add_doc(db, "9999", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "9999", "operating_profit", 8_000_000_000,
                  raw_label="前回予想:営業利益")
        _add_fact(db, doc_id, "9999", "operating_profit", 10_500_000_000,
                  raw_label="今回予想:営業利益")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["delta_value"] == 2_500_000_000
        assert abs(r["delta_pct"] - 0.3125) < 0.001

    def test_before_value_zero_safe(self, db):
        """before_value = 0 の場合 delta_pct は None"""
        doc_id = _add_doc(db, "1234", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "1234", "net_income", 0,
                  raw_label="前回予想:当期純利益")
        _add_fact(db, doc_id, "1234", "net_income", 500_000_000,
                  raw_label="今回予想:当期純利益")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["before_value"] == 0
        assert r["after_value"] == 500_000_000
        assert r["delta_value"] == 500_000_000
        assert r["delta_pct"] is None

    def test_single_value_only(self, db):
        """片方しかない場合、after_value のみ埋まる"""
        doc_id = _add_doc(db, "5678", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "5678", "sales", 50_000_000_000,
                  raw_label="修正後:売上高")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["before_value"] is None
        assert r["after_value"] == 50_000_000_000
        assert r["delta_value"] is None

    def test_three_or_more_skip(self, db):
        """同一キーで3件以上の場合、ペアリングしない（skip）"""
        doc_id = _add_doc(db, "7777", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "7777", "sales", 100_000_000,
                  raw_label="1回目:売上高")
        _add_fact(db, doc_id, "7777", "sales", 200_000_000,
                  raw_label="2回目:売上高")
        _add_fact(db, doc_id, "7777", "sales", 300_000_000,
                  raw_label="3回目:売上高")

        rows = build_forecast_revision_rows(db)
        # 3件は skip されるので0件
        assert len(rows) == 0

    def test_low_confidence_excluded(self, db):
        """confidence=low は除外"""
        doc_id = _add_doc(db, "8888", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "8888", "sales", 100_000_000,
                  raw_label="前回予想", confidence="low")
        _add_fact(db, doc_id, "8888", "sales", 200_000_000,
                  raw_label="今回予想", confidence="low")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 0

    def test_id_order_fallback(self, db):
        """raw_label にキーワードがない場合は id 小=before, 大=after"""
        doc_id = _add_doc(db, "3333", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "3333", "sales", 80_000_000_000,
                  raw_label="行1:売上高")
        _add_fact(db, doc_id, "3333", "sales", 95_000_000_000,
                  raw_label="行2:売上高")

        rows = build_forecast_revision_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["before_value"] == 80_000_000_000
        assert r["after_value"] == 95_000_000_000


# ============================================================
# MONTHLY_DATA Tests
# ============================================================

class TestMonthlyData:

    def test_year_month_extraction(self, db):
        """period から year_month が取れる"""
        doc_id = _add_doc(db, "2670", "2027-02-10", "月次売上", "monthly")
        _add_fact(db, doc_id, "2670", "monthly_sales", 5_000_000_000,
                  period="2027-01-31")

        rows = build_monthly_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["year_month"] == "2027-01"
        assert rows[0]["pubdate"] == "2027-02-10"

    def test_year_month_missing_keeps_pubdate(self, db):
        """year_month が取れなくても pubdate を保持して出力"""
        doc_id = _add_doc(db, "2670", "2027-02-10", "月次", "monthly")
        _add_fact(db, doc_id, "2670", "monthly_sales", 3_000_000_000,
                  period="")

        rows = build_monthly_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["year_month"] == ""  # 取れない
        assert rows[0]["pubdate"] == "2027-02-10"  # 保持

    def test_monthly_metrics(self, db):
        """monthly 系メトリックが正しく分類される"""
        doc_id = _add_doc(db, "3099", "2027-03-01", "月次", "monthly")
        _add_fact(db, doc_id, "3099", "same_store_sales_yoy", 105.2,
                  unit="%", period="2027-02-28")
        _add_fact(db, doc_id, "3099", "customer_count", 150000,
                  period="2027-02-28")

        rows = build_monthly_data_rows(db)
        assert len(rows) == 2
        metrics = {r["metric_name"] for r in rows}
        assert "same_store_sales_yoy" in metrics
        assert "customer_count" in metrics

    def test_doc_type_column(self, db):
        """doc_type 列が保持される"""
        doc_id = _add_doc(db, "2670", "2027-02-10", "月次", "monthly")
        _add_fact(db, doc_id, "2670", "monthly_sales", 1_000_000,
                  period="2027-01-31")

        rows = build_monthly_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["doc_type"] == "monthly"


# ============================================================
# KPI_DATA Tests
# ============================================================

class TestKpiData:

    def test_presentation_included(self, db):
        """presentation doc_type が出力される"""
        doc_id = _add_doc(db, "4062", "2027-02-10", "決算説明資料", "presentation",
                          file_type="pdf")
        _add_fact(db, doc_id, "4062", "arpu", 1500,
                  source_type="pdf")

        rows = build_kpi_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["metric_name"] == "arpu"
        assert rows[0]["doc_type"] == "presentation"

    def test_supplement_kpi_segment_included(self, db):
        """supplement/kpi/segment の doc_type が出力される"""
        for dt in ("supplement", "kpi", "segment"):
            doc_id = _add_doc(db, "1234", "2027-02-10", f"{dt}資料", dt)
            _add_fact(db, doc_id, "1234", "store_count", 100 + hash(dt) % 100)

        rows = build_kpi_data_rows(db)
        doc_types = {r["doc_type"] for r in rows}
        assert "supplement" in doc_types
        assert "kpi" in doc_types
        assert "segment" in doc_types

    def test_low_confidence_excluded(self, db):
        """confidence=low は除外"""
        doc_id = _add_doc(db, "5555", "2027-02-10", "KPI", "kpi")
        _add_fact(db, doc_id, "5555", "utilization_rate", 65.0,
                  confidence="low")
        _add_fact(db, doc_id, "5555", "store_count", 200,
                  confidence="high")

        rows = build_kpi_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["metric_name"] == "store_count"

    def test_preserves_segment_page_table(self, db):
        """segment_name / page_no / table_title が保持される"""
        doc_id = _add_doc(db, "6666", "2027-02-10", "セグメント", "segment",
                          file_type="pdf")
        _add_fact(db, doc_id, "6666", "segment_sales", 5_000_000_000,
                  segment_name="ゲーム事業", table_title="セグメント別売上",
                  page_no=3, source_type="pdf")

        rows = build_kpi_data_rows(db)
        assert len(rows) == 1
        r = rows[0]
        assert r["segment_name"] == "ゲーム事業"
        assert r["page_no"] == 3
        assert r["table_title"] == "セグメント別売上"

    def test_forecast_only_metrics_excluded(self, db):
        """forecast_revision 専用メトリックは KPI から除外"""
        doc_id = _add_doc(db, "7777", "2027-02-10", "補足", "supplement")
        _add_fact(db, doc_id, "7777", "forecast_sales", 10_000_000_000)
        _add_fact(db, doc_id, "7777", "orders", 2_000_000_000)

        rows = build_kpi_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["metric_name"] == "orders"

    def test_doc_type_column(self, db):
        """doc_type 列が保持される"""
        doc_id = _add_doc(db, "4062", "2027-02-10", "KPI一覧", "kpi")
        _add_fact(db, doc_id, "4062", "arpu", 1200)

        rows = build_kpi_data_rows(db)
        assert len(rows) == 1
        assert rows[0]["doc_type"] == "kpi"


# ============================================================
# 共通 Tests
# ============================================================

class TestCommon:

    def test_ticker_is_string(self, db):
        """ticker が文字列型で出る"""
        doc_id = _add_doc(db, "4062", "2027-02-10", "月次", "monthly")
        _add_fact(db, doc_id, "4062", "monthly_sales", 1_000_000)

        rows = build_monthly_data_rows(db)
        assert isinstance(rows[0]["ticker"], str)

    def test_value_in_yen(self, db):
        """金額が円単位のまま出る"""
        doc_id = _add_doc(db, "9999", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc_id, "9999", "sales", 120_000_000_000,
                  unit="円", raw_label="今回予想:売上高")

        rows = build_forecast_revision_rows(db)
        assert rows[0]["after_value"] == 120_000_000_000

    def test_write_sheet_creates_sheet(self, db):
        """write_sheet でシートが作成される"""
        import openpyxl
        wb = openpyxl.Workbook()
        rows = [{"ticker": "4062", "metric_name": "arpu", "metric_value": 1500}]
        cols = ["ticker", "metric_name", "metric_value"]
        count = write_sheet(wb, "TEST_SHEET", rows, cols, "2027-01-01 12:00:00")
        assert count == 1
        assert "TEST_SHEET" in wb.sheetnames
        ws = wb["TEST_SHEET"]
        assert ws.cell(row=2, column=1).value == "ticker"
        assert ws.cell(row=3, column=1).value == "4062"

    def test_write_sheet_overwrites_existing(self, db):
        """同名シートが既に存在する場合は削除して再作成"""
        import openpyxl
        wb = openpyxl.Workbook()
        # 最初の作成
        rows1 = [{"a": "old"}]
        write_sheet(wb, "MY_SHEET", rows1, ["a"])
        # 上書き
        rows2 = [{"a": "new1"}, {"a": "new2"}]
        count = write_sheet(wb, "MY_SHEET", rows2, ["a"])
        assert count == 2
        ws = wb["MY_SHEET"]
        assert ws.cell(row=3, column=1).value == "new1"
        assert ws.cell(row=4, column=1).value == "new2"

    def test_write_extracted_facts_sheets_no_tables(self):
        """documents/extracted_facts テーブルが無い場合もエラーにならない"""
        import openpyxl
        conn = sqlite3.connect(":memory:")
        wb = openpyxl.Workbook()
        stats = write_extracted_facts_sheets(wb, conn)
        assert stats["forecast_rows"] == 0
        assert stats["monthly_rows"] == 0
        assert stats["kpi_rows"] == 0
        conn.close()

    def test_write_extracted_facts_sheets_creates_three_sheets(self, db):
        """3シートが正しく作成される"""
        import openpyxl
        wb = openpyxl.Workbook()

        # FORECAST
        doc1 = _add_doc(db, "4062", "2027-01-15", "修正", "forecast_revision")
        _add_fact(db, doc1, "4062", "sales", 100e9, raw_label="前回予想:売上高")
        _add_fact(db, doc1, "4062", "sales", 120e9, raw_label="今回予想:売上高")

        # MONTHLY
        doc2 = _add_doc(db, "2670", "2027-02-10", "月次", "monthly")
        _add_fact(db, doc2, "2670", "monthly_sales", 5e9, period="2027-01-31")

        # KPI
        doc3 = _add_doc(db, "4062", "2027-02-10", "説明資料", "presentation")
        _add_fact(db, doc3, "4062", "arpu", 1500, source_type="pdf")

        stats = write_extracted_facts_sheets(wb, db, "2027-03-08 10:00:00")
        assert stats["forecast_rows"] == 1
        assert stats["monthly_rows"] == 1
        assert stats["kpi_rows"] == 1
        assert FORECAST_SHEET in wb.sheetnames
        assert MONTHLY_SHEET in wb.sheetnames
        assert KPI_SHEET in wb.sheetnames
