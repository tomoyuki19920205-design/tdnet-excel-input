#!/usr/bin/env python3
# ============================================================
# xlsx_stats.py — data.xlsx の統計を出力
# ============================================================
# run_update_all.bat の Step4 で呼ばれる。
# コンソール + ログに行数・ticker数・前回差分を出力。
# ============================================================
from __future__ import annotations

import argparse
import io
import json
import os
import shutil
import sys
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JST = timezone(timedelta(hours=9))
_STATS_FILE = "data/xlsx_prev_stats.json"


def _load_prev(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_stats(path: str, data: dict) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    data["saved_at"] = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def xlsx_stats(xlsx_path: str, stats_path: str = "") -> dict:
    """data.xlsx を読み込み、統計を返す"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[XLSX] openpyxl がインストールされていません")
        return {}

    if not os.path.exists(xlsx_path):
        print(f"[XLSX] ファイルが見つかりません: {xlsx_path}")
        return {}

    if not stats_path:
        stats_path = os.path.join(_PROJECT_ROOT, _STATS_FILE)

    # OneDrive ロック対策: 一時コピーで読む
    tmp_path = None
    read_path = xlsx_path
    try:
        wb = load_workbook(
            read_path, read_only=True, data_only=True
        )
    except PermissionError:
        # shutil.copy2 → robocopy の順でフォールバック
        import subprocess
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        copied = False

        # 方法1: shutil.copy2
        try:
            shutil.copy2(xlsx_path, tmp_path)
            copied = True
        except PermissionError:
            pass

        # 方法2: robocopy (Windows)
        if not copied:
            try:
                src_dir = os.path.dirname(xlsx_path)
                src_name = os.path.basename(xlsx_path)
                tmp_dir = os.path.dirname(tmp_path)
                tmp_name = os.path.basename(tmp_path)
                subprocess.run(
                    ["robocopy", src_dir, tmp_dir,
                     src_name, "/IS", "/IT"],
                    capture_output=True, timeout=10,
                )
                # robocopy は出力ファイル名を元のまま使う
                robo_out = os.path.join(tmp_dir, src_name)
                if os.path.exists(robo_out):
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                    os.rename(robo_out, tmp_path)
                    copied = True
            except Exception:
                pass

        if not copied:
            # 完全に読めない場合: 前回統計だけ表示
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            print(
                "[XLSX] ファイルがロック中 "
                "(コピーも不可) → 前回統計のみ表示"
            )
            prev = _load_prev(stats_path)
            if prev:
                print(
                    f"[XLSX] 前回: rows={prev.get('rows',0)} "
                    f"tickers={prev.get('tickers',0)}"
                )
            else:
                print("[XLSX] 前回統計なし")
            return prev

        print(
            "[XLSX] ファイルがロック中 → "
            "一時コピーで読み込み"
        )
        read_path = tmp_path
        wb = load_workbook(
            read_path, read_only=True, data_only=True
        )

    try:
        # DATA シートを探す
        if "DATA" in wb.sheetnames:
            ws = wb["DATA"]
        else:
            ws = wb.active

        # 行数カウント (ヘッダー除外)
        rows = 0
        tickers: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows += 1
            if row[0] is not None:
                tickers.add(str(row[0]))
    finally:
        wb.close()
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)

    now_stats = {"rows": rows, "tickers": len(tickers)}

    # 前回との差分
    prev = _load_prev(stats_path)
    prev_rows = prev.get("rows", 0)
    delta = rows - prev_rows

    # 出力
    print(f"[XLSX] rows(DATA)={rows}")
    if prev_rows > 0:
        sign = "+" if delta >= 0 else ""
        print(
            f"[XLSX] delta={sign}{delta} "
            f"(prev={prev_rows}, now={rows})"
        )
    else:
        print("[XLSX] delta=初回実行 (前回データなし)")
    print(f"[XLSX] tickers={len(tickers)}")

    # 今回の統計を保存
    _save_stats(stats_path, now_stats)

    return {"rows": rows, "tickers": len(tickers), "delta": delta}


def main():
    if sys.stdout and hasattr(sys.stdout, "encoding"):
        if sys.stdout.encoding and sys.stdout.encoding.lower() not in (
            "utf-8", "utf8"
        ):
            sys.stdout = io.TextIOWrapper(
                sys.stdout.buffer, encoding="utf-8",
                errors="replace",
            )

    parser = argparse.ArgumentParser(
        description="data.xlsx の統計を出力"
    )
    parser.add_argument(
        "--file", required=True,
        help="data.xlsx のパス",
    )
    args = parser.parse_args()

    xlsx_stats(args.file)


if __name__ == "__main__":
    main()
