#!/usr/bin/env python3
# ============================================================
# build_viewer_extended_views.py
# ============================================================
"""
viewer.xlsx に FORECAST_VIEW / MONTHLY_VIEW / KPI_VIEW を追加する。

既存の PL_VIEW / P シートは一切変更しない。
data.xlsx 外部参照 ('[data.xlsx]SHEET'!...) を使用する。

CLI:
  .\.venv\Scripts\python.exe tools\build_viewer_extended_views.py
  .\.venv\Scripts\python.exe tools\build_viewer_extended_views.py --output viewer_extended.xlsx
"""
from __future__ import annotations

import argparse
import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

_PROJECT_ROOT = str(Path(__file__).resolve().parent.parent)
sys.path.insert(0, _PROJECT_ROOT)

logger = logging.getLogger("viewer_ext")

# ============================================================
# 定数
# ============================================================

# 入力 viewer パス候補
_VIEWER_PATHS = [
    r"C:\Users\takuy\OneDrive\viewer_extlink_v8.xlsx",
    r"C:\Users\takuy\OneDrive\viewer.xlsx",
]

# ticker 入力セル (PL_VIEW!B2 を参照)
_TICKER_REF = "PL_VIEW!$B$2"

# data.xlsx ファイル名 (外部参照用)
_DATA_FILE = "data.xlsx"


def _ext_ref(sheet_name: str) -> str:
    """外部参照プレフィクスを生成: '[data.xlsx]SHEET'!"""
    return f"'[{_DATA_FILE}]{sheet_name}'!"

# スタイル
_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=10, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_LABEL_FONT = Font(bold=True, size=11)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)
_PCT_FORMAT = '0.0%'
_NUM_FORMAT = '#,##0'
_DATE_FORMAT = 'yyyy-mm-dd'

# 各シート data.xlsx 列マッピング (1-indexed, data行は3行目から)
# FORECAST_REVISION: A=ticker, B=pubdate, C=title, D=doc_type, E=period, F=quarter,
#   G=metric_name, H=before_value, I=after_value, J=delta_value, K=delta_pct,
#   L=unit, M=source_type, N=confidence, O=document_id, P=raw_label, Q=normalized_label

_FORECAST_SRC_SHEET = "FORECAST_REVISION"
_FORECAST_DISPLAY_COLS = [
    # (header, src_col_letter, format, width)
    ("pubdate",       "B", None,         12),
    ("title",         "C", None,         30),
    ("period",        "E", None,         12),
    ("quarter",       "F", None,         8),
    ("metric_name",   "G", None,         18),
    ("before_value",  "H", _NUM_FORMAT,  16),
    ("after_value",   "I", _NUM_FORMAT,  16),
    ("delta_value",   "J", _NUM_FORMAT,  16),
    ("delta_pct",     "K", _PCT_FORMAT,  10),
    ("confidence",    "N", None,         10),
    ("source_type",   "M", None,         10),
]

# MONTHLY_DATA: A=ticker, B=pubdate, C=title, D=doc_type, E=year_month,
#   F=metric_name, G=metric_value, H=unit, I=segment_name, J=source_type,
#   K=confidence, L=document_id, M=raw_label, N=normalized_label

_MONTHLY_SRC_SHEET = "MONTHLY_DATA"
_MONTHLY_DISPLAY_COLS = [
    ("pubdate",       "B", None,         12),
    ("title",         "C", None,         30),
    ("year_month",    "E", None,         12),
    ("metric_name",   "F", None,         18),
    ("metric_value",  "G", _NUM_FORMAT,  16),
    ("unit",          "H", None,         8),
    ("segment_name",  "I", None,         16),
    ("confidence",    "K", None,         10),
    ("source_type",   "J", None,         10),
]

# KPI_DATA: A=ticker, B=pubdate, C=title, D=doc_type, E=period, F=quarter,
#   G=metric_name, H=metric_value, I=unit, J=segment_name, K=source_type,
#   L=confidence, M=document_id, N=table_title, O=page_no,
#   P=raw_label, Q=normalized_label

_KPI_SRC_SHEET = "KPI_DATA"
_KPI_DISPLAY_COLS = [
    ("pubdate",       "B", None,         12),
    ("title",         "C", None,         28),
    ("period",        "E", None,         12),
    ("quarter",       "F", None,         8),
    ("metric_name",   "G", None,         18),
    ("metric_value",  "H", _NUM_FORMAT,  16),
    ("unit",          "I", None,         8),
    ("segment_name",  "J", None,         16),
    ("table_title",   "N", None,         24),
    ("page_no",       "O", None,         8),
    ("confidence",    "L", None,         10),
    ("source_type",   "K", None,         10),
]


# ============================================================
# シート構築
# ============================================================

def _build_view_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    title: str,
    src_sheet: str,
    display_cols: list[tuple],
    max_rows: int,
):
    """
    MATCH + INDEX 方式の VIEW シートを構築する。
    '[data.xlsx]SHEET'! 形式の外部参照を使用する。
    """
    ext = _ext_ref(src_sheet)  # e.g. "'[data.xlsx]FORECAST_REVISION'!"
    # 既存シート削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # --- ヘッダーエリア ---
    # A1: タイトル
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT

    # A2: "ticker" ラベル、B2: ticker 参照
    ws["A2"] = "ticker ▶"
    ws["A2"].font = _LABEL_FONT
    ws["B2"] = f"={_TICKER_REF}"
    ws["B2"].font = Font(bold=True, size=14, color="FF0000")

    # C2: TEXT変換 (4桁→文字列)
    ws["C2"] = f'=TEXT(B2,"0")'
    ws["C2"].font = Font(color="999999", size=9)

    # --- MATCH/COUNTIF (隠し計算セル) ---
    # 行28を使用 (PL_VIEW と同じパターン)
    match_row = 28 + max_rows  # データ行の下に配置
    ticker_col = "A"  # FORECAST_REVISION/MONTHLY_DATA/KPI_DATA の A列 = ticker

    # 開始行 (MATCH)
    ws.cell(row=match_row, column=1,
            value=f'=IFERROR(MATCH(TEXT($B$2,"0"),{ext}$A$3:$A$200000,0),0)')
    ws.cell(row=match_row, column=1).font = Font(color="999999", size=8)

    # ヒット数 (COUNTIF)
    ws.cell(row=match_row, column=2,
            value=f'=COUNTIF({ext}$A$3:$A$200000,TEXT($B$2,"0"))')
    ws.cell(row=match_row, column=2).font = Font(color="999999", size=8)

    # ラベル
    match_ref = f"$A${match_row}"
    count_ref = f"$B${match_row}"

    # --- ヘッダー行 (4行目) ---
    header_row = 4
    for col_idx, (header, src_col, fmt, width) in enumerate(display_cols, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        # 列幅
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # --- データ行 (5行目から) ---
    data_start = 5
    for row_offset in range(max_rows):
        row_num = data_start + row_offset
        for col_idx, (header, src_col, fmt, width) in enumerate(display_cols, 1):
            # IF(OR(MATCH=0, offset >= MIN(COUNT, max)), "",
            #    INDEX([1]SHEET!$col$3:$col$200000, MATCH+offset))
            formula = (
                f'=IF(OR({match_ref}=0,{row_offset}>='
                f'MIN({count_ref},{max_rows})),"",'
                f'INDEX({ext}${src_col}$3:${src_col}$200000,'
                f'{match_ref}+{row_offset}))'
            )
            cell = ws.cell(row=row_num, column=col_idx, value=formula)
            if fmt:
                cell.number_format = fmt
            cell.border = _THIN_BORDER

    # --- 該当なし表示 ---
    no_data_row = data_start + max_rows + 1
    ws.cell(row=no_data_row, column=1,
            value=f'=IF({match_ref}=0,"該当データなし","")')
    ws.cell(row=no_data_row, column=1).font = Font(color="FF0000", italic=True)

    # --- フリーズペイン ---
    ws.freeze_panes = f"A{data_start}"

    logger.info(f"[VIEWER] {sheet_name} 構築完了 (max {max_rows} rows)")


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="viewer.xlsx に FORECAST/MONTHLY/KPI VIEW を追加"
    )
    parser.add_argument(
        "--input", default="",
        help="入力 viewer.xlsx パス (省略時は自動検出)",
    )
    parser.add_argument(
        "--output", "-o", default="",
        help="出力パス (省略時は viewer_extended.xlsx)",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="[%(asctime)s] %(message)s", datefmt="%H:%M:%S",
    )

    # --- 入力 viewer を探す ---
    input_path = args.input
    if not input_path:
        for p in _VIEWER_PATHS:
            if os.path.exists(p):
                input_path = p
                break
    if not input_path or not os.path.exists(input_path):
        print("ERROR: viewer.xlsx が見つかりません")
        sys.exit(1)

    # --- 出力パス ---
    output_path = args.output
    if not output_path:
        output_dir = os.path.dirname(input_path)
        output_path = os.path.join(output_dir, "viewer_extended.xlsx")

    print()
    print("=" * 50)
    print("  viewer 拡張ツール")
    print("=" * 50)
    print(f"  入力: {input_path}")
    print(f"  出力: {output_path}")
    print()

    # --- viewer を開く ---
    logger.info(f"[VIEWER] Opening: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    existing_sheets = wb.sheetnames[:]
    logger.info(f"[VIEWER] 既存シート: {existing_sheets}")

    # --- 既存シート確認 ---
    if "PL_VIEW" not in existing_sheets:
        print("WARNING: PL_VIEW が見つかりません。ticker 参照が動かない可能性があります。")

    # --- 3シート構築 ---
    _build_view_sheet(
        wb, "FORECAST_VIEW", "[FORECAST] gyouseki yosou shuusei",
        _FORECAST_SRC_SHEET, _FORECAST_DISPLAY_COLS, max_rows=10,
    )

    _build_view_sheet(
        wb, "MONTHLY_VIEW", "[MONTHLY] getsuji data",
        _MONTHLY_SRC_SHEET, _MONTHLY_DISPLAY_COLS, max_rows=20,
    )

    _build_view_sheet(
        wb, "KPI_VIEW", "[KPI] hosoku data",
        _KPI_SRC_SHEET, _KPI_DISPLAY_COLS, max_rows=30,
    )

    # --- 保存 ---
    wb.save(output_path)
    logger.info(f"[VIEWER] 保存完了: {output_path}")

    print()
    print("=" * 50)
    print("  [OK] 完了")
    print("=" * 50)
    print(f"  出力: {output_path}")
    print(f"  シート: {wb.sheetnames}")
    print()
    print("  確認手順:")
    print(f"  1. {output_path} を Excel で開く")
    print("  2. PL_VIEW!B2 に銘柄コード (例: 4062) を入力")
    print("  3. FORECAST_VIEW / MONTHLY_VIEW / KPI_VIEW を確認")
    print()
    print("  注意: 外部リンク更新を求められたら「更新」を選択してください")
    print("=" * 50)
    print()

    wb.close()


if __name__ == "__main__":
    main()
