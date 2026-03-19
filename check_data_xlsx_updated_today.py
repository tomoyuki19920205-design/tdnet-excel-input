from openpyxl import load_workbook

PATH = r"C:\Users\takuy\OneDrive\data.xlsx"
wb = load_workbook(PATH, data_only=True, read_only=True)
ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active

# ヘッダー行探す（1〜50）
header_row = None
header = None
for r in range(1, 51):
    vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
    keys = set([str(v).strip().lower() for v in vals if v is not None])
    if "ticker" in keys and "updated_at" in keys:
        header_row = r
        header = vals
        break

print("header_row=", header_row)
print("header=", header)

hmap = {str(v).strip().lower(): i for i,v in enumerate(header) if v is not None}
i_t = hmap["ticker"]
i_u = hmap["updated_at"]

# updated_at に "2026-03-04" を含む行を上位50件
hits = []
for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    ua = row[i_u]
    if ua is None:
        continue
    s = str(ua)
    if "2026-03-04" in s:
        hits.append((row[i_t], s))
        if len(hits) >= 50:
            break

print("\n=== data.xlsx rows updated_at contains 2026-03-04 (top 50) ===")
print("count=", len(hits))
for x in hits:
    print(x)

wb.close()
