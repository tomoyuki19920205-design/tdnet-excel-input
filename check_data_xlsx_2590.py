from openpyxl import load_workbook

PATH = r"C:\Users\takuy\OneDrive\data.xlsx"
wb = load_workbook(PATH, data_only=True, read_only=True)
ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active

# 1〜50行で、ticker/company_code/code/period/quarter/sales っぽい列名がある行をヘッダーとみなす
header_row = None
header = None
for r in range(1, 51):
    vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
    norm = [str(v).strip().lower() for v in vals if v is not None]
    keys = set(norm)
    if ("ticker" in keys) or ("company_code" in keys) or ("code" in keys):
        header_row = r
        header = vals
        break

print("header_row=", header_row)
print("header=", header)

if header_row is None:
    raise SystemExit("FATAL: ヘッダー行が見つからない（1〜50行を確認）")

hmap = {}
for i,v in enumerate(header):
    if v is None: 
        continue
    hmap[str(v).strip().lower()] = i

def find_col(*cands):
    for c in cands:
        c = c.lower()
        if c in hmap:
            return hmap[c]
    return None

i_code   = find_col("ticker","company_code","code")
i_period = find_col("period","fiscal_year_end","fy_end")
i_q      = find_col("quarter","q")
i_sales  = find_col("sales","revenue")
i_op     = find_col("operating_profit","op")

print("cols=", {"code":i_code,"period":i_period,"q":i_q,"sales":i_sales,"op":i_op})

targets = {"2590","259"}
hits = {t:[] for t in targets}

for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    code = row[i_code] if i_code is not None else None
    if code is None:
        continue
    code_s = str(code).strip()
    if code_s in targets:
        rec = (
            code_s,
            row[i_period] if i_period is not None else None,
            row[i_q] if i_q is not None else None,
            row[i_sales] if i_sales is not None else None,
            row[i_op] if i_op is not None else None,
        )
        if len(hits[code_s]) < 50:
            hits[code_s].append(rec)

for t in ["2590","259"]:
    print("\n=== data.xlsx hits:", t, "count_shown=", len(hits[t]), "===")
    for r in hits[t]:
        print(r)

wb.close()
