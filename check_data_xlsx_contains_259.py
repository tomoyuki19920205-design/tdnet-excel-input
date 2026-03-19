from openpyxl import load_workbook

PATH = r"C:\Users\takuy\OneDrive\data.xlsx"
wb = load_workbook(PATH, data_only=True, read_only=True)
ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active

# ヘッダー行を探す（1〜50）
header_row = None
header = None
for r in range(1, 51):
    vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
    keys = set([str(v).strip().lower() for v in vals if v is not None])
    if "ticker" in keys:
        header_row = r
        header = vals
        break

print("header_row=", header_row)
print("header=", header)

if header_row is None:
    raise SystemExit("FATAL: header row not found")

hmap = {}
for i,v in enumerate(header):
    if v is None:
        continue
    hmap[str(v).strip().lower()] = i

i_code = hmap.get("ticker", None)
i_period = hmap.get("period", None)
i_q = hmap.get("quarter", None)
i_sales = hmap.get("sales", None)
i_op = hmap.get("operating_profit", None)

print("cols=", {"ticker":i_code,"period":i_period,"q":i_q,"sales":i_sales,"op":i_op})

def norm_code(x):
    if x is None:
        return ""
    s = str(x).strip()
    # "2590.0" みたいなのを "2590" に寄せる
    if s.endswith(".0"):
        s = s[:-2]
    return s

# 259 を含むものを上位50件表示（2590/259/2590.0 など拾う）
hits = []
for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    code = norm_code(row[i_code])
    if "259" in code:  # 部分一致
        rec = (code,
               row[i_period] if i_period is not None else None,
               row[i_q] if i_q is not None else None,
               row[i_sales] if i_sales is not None else None,
               row[i_op] if i_op is not None else None)
        hits.append(rec)
        if len(hits) >= 50:
            break

print("\n=== data.xlsx rows where ticker contains '259' (top 50) ===")
print("count=", len(hits))
for r in hits:
    print(r)

wb.close()
