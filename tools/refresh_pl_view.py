#!/usr/bin/env python3
# ============================================================
# refresh_pl_view.py
# ============================================================
# data.xlsx の DATA シートを読み取り、
# ビューア Excel の _DATA シートを差し替える。
# PL_VIEW シートの数式やレイアウトは一切触らない。
#
# 使い方:
#   python tools\refresh_pl_view.py
#   python tools\refresh_pl_view.py --data_xlsx "...\data.xlsx" --viewer_xlsx "...\テスト用.xlsx"
# ============================================================
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

JST = timezone(timedelta(hours=9))

# デフォルトパス
_DEFAULT_DATA = r"C:\Users\takuy\OneDrive\data.xlsx"
_DEFAULT_VIEWER = r"C:\Users\takuy\OneDrive\20260303テスト用コピー.xlsx"
_DEFAULT_SHEET = "_DATA"


def _safe_print(msg: str):
    """cp932 でも壊れない print"""
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode("ascii", "replace").decode("ascii"))


def refresh(
    data_xlsx: str,
    viewer_xlsx: str,
    sheet_name: str = _DEFAULT_SHEET,
    backup: bool = True,
) -> dict:
    """
    data_xlsx の DATA シートを読み取り、
    viewer_xlsx の sheet_name シートを丸ごと差し替える。
    PL_VIEW 等の他シートは一切触らない。
    """

    # ---- 入力ファイル存在チェック ----
    if not os.path.exists(data_xlsx):
        _safe_print(f"[ERROR] data.xlsx が見つかりません: {data_xlsx}")
        _safe_print("  run_update_all.bat を先に実行してください。")
        return {"ok": False, "error": "data.xlsx not found"}

    if not os.path.exists(viewer_xlsx):
        _safe_print(f"[ERROR] ビューア Excel が見つかりません: {viewer_xlsx}")
        return {"ok": False, "error": "viewer_xlsx not found"}

    # ---- バックアップ ----
    if backup:
        ts = datetime.now(JST).strftime("%Y%m%d_%H%M%S")
        bak_path = viewer_xlsx + f".{ts}.bak"
        try:
            shutil.copy2(viewer_xlsx, bak_path)
            _safe_print(f"[BACKUP] {os.path.basename(bak_path)}")
        except Exception as e:
            _safe_print(f"[WARN] バックアップ作成失敗 (続行します): {e}")

    # ---- data.xlsx 読み取り ----
    _safe_print(f"[1/3] data.xlsx を読み込み中...")
    try:
        data_wb = openpyxl.load_workbook(data_xlsx, data_only=True, read_only=True)
    except Exception as e:
        _safe_print(f"[ERROR] data.xlsx を開けません: {e}")
        return {"ok": False, "error": str(e)}

    # DATA シートを探す
    src_sheet_name = None
    for name in ["DATA", "data", "Sheet1"]:
        if name in data_wb.sheetnames:
            src_sheet_name = name
            break
    if not src_sheet_name:
        _safe_print(f"[ERROR] data.xlsx に DATA シートがありません")
        _safe_print(f"  存在するシート: {data_wb.sheetnames}")
        data_wb.close()
        return {"ok": False, "error": "DATA sheet not found in data.xlsx"}

    src_ws = data_wb[src_sheet_name]
    src_data = []
    for row in src_ws.iter_rows(min_row=1, values_only=True):
        src_data.append(row)
    data_wb.close()

    if not src_data:
        _safe_print("[ERROR] data.xlsx の DATA シートが空です")
        return {"ok": False, "error": "DATA sheet is empty"}

    total_rows = len(src_data)
    total_cols = max(len(r) for r in src_data) if src_data else 0
    _safe_print(f"  -> {total_rows:,} rows x {total_cols} cols")

    # ---- ビューア Excel を開く ----
    _safe_print(f"[2/3] ビューア Excel を開いています...")
    try:
        viewer_wb = openpyxl.load_workbook(viewer_xlsx)
    except Exception as e:
        _safe_print(f"[ERROR] ビューア Excel を開けません: {e}")
        _safe_print("  ファイルが Excel で開かれていたら閉じてください。")
        return {"ok": False, "error": str(e)}

    # _DATA シートの存在チェック
    if sheet_name not in viewer_wb.sheetnames:
        _safe_print(f"[ERROR] ビューア Excel に '{sheet_name}' シートがありません")
        _safe_print(f"  存在するシート: {viewer_wb.sheetnames}")
        viewer_wb.close()
        return {"ok": False, "error": f"'{sheet_name}' not found in viewer"}

    # ---- _DATA シートを差し替え ----
    _safe_print(f"[3/3] {sheet_name} シートを更新中...")

    # 位置を記録してから削除→再作成（シート順序を保持）
    sheet_idx = viewer_wb.sheetnames.index(sheet_name)
    old_tab_color = viewer_wb[sheet_name].sheet_properties.tabColor
    del viewer_wb[sheet_name]

    new_ws = viewer_wb.create_sheet(sheet_name, sheet_idx)
    if old_tab_color:
        new_ws.sheet_properties.tabColor = old_tab_color

    # データ書き込み
    for r_idx, row_data in enumerate(src_data, 1):
        for c_idx, value in enumerate(row_data, 1):
            new_ws.cell(row=r_idx, column=c_idx, value=value)

    # ヘッダー行（2行目）を太字にする
    if total_rows >= 2:
        for c in range(1, total_cols + 1):
            cell = new_ws.cell(row=2, column=c)
            cell.font = Font(bold=True)

    # 列幅をそれなりに
    col_widths = [12, 14, 10, 18, 18, 20, 10, 26]
    for i, w in enumerate(col_widths):
        if i < total_cols:
            from openpyxl.utils import get_column_letter
            new_ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ---- PL_VIEW の FILTER 数式を更新（行数が変わった場合） ----
    if "PL_VIEW" in viewer_wb.sheetnames:
        pl = viewer_wb["PL_VIEW"]
        a5_val = pl["A5"].value
        if a5_val and isinstance(a5_val, str) and "FILTER" in a5_val:
            # 行数を新しいデータに合わせて更新
            formula = (
                '=_xlfn._xlws.FILTER('
                f'{sheet_name}!A$3:F${total_rows},'
                f'LEFT({sheet_name}!A$3:A${total_rows},4)='
                f'LEFT(_xlfn.TEXT($B$2,"0"),4),'
                '"該当なし")'
            )
            pl["A5"] = formula
            _safe_print(f"  PL_VIEW!A5 の FILTER 数式を更新 (max_row={total_rows})")

    # ---- 保存 ----
    try:
        viewer_wb.save(viewer_xlsx)
    except PermissionError:
        _safe_print("[ERROR] 保存できません。ビューア Excel を閉じてから再実行してください。")
        viewer_wb.close()
        return {"ok": False, "error": "PermissionError"}
    except Exception as e:
        _safe_print(f"[ERROR] 保存に失敗しました: {e}")
        viewer_wb.close()
        return {"ok": False, "error": str(e)}

    viewer_wb.close()

    _safe_print(f"REFRESH_OK rows={total_rows} cols={total_cols}")
    return {"ok": True, "rows": total_rows, "cols": total_cols}


def main():
    parser = argparse.ArgumentParser(
        description="data.xlsx の内容でビューア Excel の _DATA を更新する",
    )
    parser.add_argument(
        "--data_xlsx", default=_DEFAULT_DATA,
        help=f"データ元 (default: {_DEFAULT_DATA})",
    )
    parser.add_argument(
        "--viewer_xlsx", default=_DEFAULT_VIEWER,
        help=f"更新先ビューア (default: {_DEFAULT_VIEWER})",
    )
    parser.add_argument(
        "--sheet", default=_DEFAULT_SHEET,
        help=f"更新するシート名 (default: {_DEFAULT_SHEET})",
    )
    parser.add_argument(
        "--no-backup", action="store_true",
        help="バックアップを作らない",
    )
    args = parser.parse_args()

    result = refresh(
        data_xlsx=args.data_xlsx,
        viewer_xlsx=args.viewer_xlsx,
        sheet_name=args.sheet,
        backup=not args.no_backup,
    )

    sys.exit(0 if result.get("ok") else 1)


if __name__ == "__main__":
    main()
