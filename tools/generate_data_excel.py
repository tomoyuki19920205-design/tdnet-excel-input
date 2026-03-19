#!/usr/bin/env python3
# ============================================================
# generate_data_excel.py
# ============================================================
# Supabase public.financials → data.xlsx (フラットテーブル)
#
# ダブルクリック実行:
#   run_generate_data.bat
#
# コマンド実行:
#   .\.venv\Scripts\python.exe -m tools.generate_data_excel
#   .\.venv\Scripts\python.exe -m tools.generate_data_excel --output "C:\Users\takuy\OneDrive\data.xlsx"
#
# データソース:
#   Supabase public.financials (72k+ rows, 3800+ tickers)
#   ※ 旧方式の companies/periods/facts からの取得は廃止
# ============================================================
from __future__ import annotations

import argparse
import calendar
import logging
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import time
import traceback
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import requests

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from src.common_ticker import normalize_ticker as _normalize_ticker
from src.extraction.extracted_facts_sheets import write_extracted_facts_sheets

logger = logging.getLogger("gen_data")

JST = timezone(timedelta(hours=9))

# ============================================================
# 定数
# ============================================================
SHEET_NAME = "DATA"
HEADERS = [
    "ticker",
    "period",
    "quarter",
    "sales",
    "gross_profit",
    "operating_profit",
    "source",
    "updated_at",
    "recency_key",
    "lookup_key",
]

SEGMENT_SHEET_NAME = "SEGMENT_DATA"
SEGMENT_HEADERS = [
    "ticker",
    "period",
    "quarter",
    "segment_name",
    "metric_name",
    "value_mil",
    "recency_key",
    "lookup_key",
]

_PAGE_SIZE = 1000  # Supabase 1リクエストあたりの上限
_RETRY_MAX = 5
_RETRY_BASE_SEC = 2.0
_DEFAULT_SQLITE_DB = "decision_db.db"


# ticker 正規化は src.common_ticker.normalize_ticker を使用
# （上記 import で _normalize_ticker として取り込み済み）


_Q_ORDER = {"FY": 5, "4Q": 4, "3Q": 3, "2Q": 2, "1Q": 1}
_Q_NUM = {"1Q": "01", "2Q": "02", "3Q": "03", "4Q": "04", "FY": "04"}

def _quarter_sort_key(q: str) -> int:
    """quarter 文字列 → ソートキー (FY=5, 4Q=4, ..., 1Q=1)"""
    return _Q_ORDER.get(q, 0)


def _make_recency_key(period: str, quarter: str) -> str:
    """period '2025-03-31' + quarter '3Q' → recency_key 'YYYYQQ' (例: '202503').
    period は ISO 日付 or YYYY-MM-DD。年と Q番号を使って生成。"""
    try:
        year = period[:4]
        qq = _Q_NUM.get(quarter, "00")
        return f"{year}{qq}"
    except Exception:
        return ""


def _make_lookup_key(ticker: str, period: str, quarter: str) -> str:
    """lookup_key 'ticker|period|quarter' (例: '7203|2025-03-31|4Q')"""
    return f"{ticker}|{period}|{quarter}"


# ============================================================
# SEGMENT_DATA 用ヘルパー
# ============================================================
_FYE_RE = re.compile(r"^(\d{4})/(\d{1,2})$")  # "2025/3" → "2025-03-31"

# VIEW 再作成用SQL
_CLEAN_VIEW_SQL = """
DROP VIEW IF EXISTS segment_financials_clean;
CREATE VIEW segment_financials_clean AS
SELECT *
FROM segment_financials
WHERE segment_name IS NOT NULL
  AND segment_name != ''
  AND length(segment_name) > 1
  AND segment_name NOT IN ('売上','利益','#VALUE!','0','月次売上','累計','ＧＰ')
  AND segment_name NOT LIKE 'UNKNOWN_%'
  AND (segment_sales IS NOT NULL OR segment_profit IS NOT NULL)
  AND (segment_sales != 0 OR segment_profit != 0);
"""


def _fye_to_iso(fye: str) -> str:
    """fiscal_year_end '2025/3' → ISO日付 '2025-03-31' に変換"""
    m = _FYE_RE.match(str(fye).strip())
    if not m:
        return str(fye)
    year, month = int(m.group(1)), int(m.group(2))
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-{last_day:02d}"


def _quarter_label(q: str) -> str:
    """quarter正規化: '1Q'→'1Q', '1'→'1Q', 1→'1Q'"""
    s = str(q).strip().upper()
    if s in ("1", "2", "3", "4"):
        return f"{s}Q"
    if s in ("1Q", "2Q", "3Q", "4Q"):
        return s
    return s


def _ensure_segment_financials_clean_view(conn: sqlite3.Connection) -> None:
    """segment_financials_clean VIEW を再作成する。"""
    objects = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type IN ('table','view')"
    ).fetchall()}
    if "segment_financials" not in objects:
        logger.warning("[SEG] segment_financials テーブルが存在しません")
        return
    conn.executescript(_CLEAN_VIEW_SQL)
    logger.info("[SEG] segment_financials_clean VIEW を再作成しました")


def _read_segment_data(db_path: str) -> list[list]:
    """decision_db.db の segment_financials_clean VIEW から SEGMENT_DATA を生成する。"""
    if not os.path.exists(db_path):
        logger.warning(f"[SEG] SQLite DB が見つかりません: {db_path} — SEGMENT_DATAスキップ")
        return []

    conn = sqlite3.connect(db_path, timeout=10)
    try:
        objects = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type IN ('table','view')"
        ).fetchall()}
        if "segment_financials_clean" not in objects:
            logger.warning("[SEG] segment_financials_clean VIEW が存在しません — スキップ")
            return []

        rows = conn.execute(
            "SELECT company_code, fiscal_year_end, quarter, "
            "       segment_name, segment_sales, segment_profit "
            "FROM segment_financials_clean "
            "ORDER BY company_code, fiscal_year_end, quarter, segment_order"
        ).fetchall()
    finally:
        conn.close()

    logger.info(f"[SEG] segment_financials_clean: {len(rows):,} 行取得")

    result: list[list] = []
    for company_code, fye, q, seg_name, seg_sales, seg_profit in rows:
        ticker = _normalize_ticker(company_code)
        period = _fye_to_iso(fye)
        quarter = _quarter_label(q)
        recency = _make_recency_key(period, quarter)
        if seg_sales is not None:
            lk = _make_lookup_key(ticker, period, quarter)
            result.append([ticker, period, quarter, seg_name, "segment_sales", seg_sales, recency, lk])
        if seg_profit is not None:
            lk = _make_lookup_key(ticker, period, quarter)
            result.append([ticker, period, quarter, seg_name, "segment_profit", seg_profit, recency, lk])

    logger.info(f"[SEG] 縦持ち変換後: {len(result):,} 行")
    return result


# ============================================================
# .env 読み込み
# ============================================================
def _load_dotenv():
    env_path = Path(_PROJECT_ROOT) / ".env"
    if not env_path.exists():
        return False
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, val = line.split("=", 1)
                os.environ.setdefault(key.strip(), val.strip())
    return True


# ============================================================
# Supabase ページネーション付き取得
# ============================================================
def _fetch_with_retry(
    url: str,
    headers: dict,
    params: dict,
    timeout: int = 30,
) -> requests.Response:
    """リトライ付きGETリクエスト (429/5xx対応)"""
    last_exc = None
    for attempt in range(_RETRY_MAX):
        try:
            r = requests.get(
                url, headers=headers,
                params=params, timeout=timeout,
            )
            r.raise_for_status()
            return r
        except (requests.ConnectionError, requests.Timeout) as e:
            last_exc = e
            wait = _RETRY_BASE_SEC * (2 ** attempt)
            logger.warning(
                f"[FETCH] 接続エラー ({attempt+1}/{_RETRY_MAX})"
                f" — {wait:.0f}秒待機"
            )
            time.sleep(wait)
        except requests.HTTPError as e:
            status = (
                e.response.status_code if e.response else 0
            )
            if status == 429 or status >= 500:
                last_exc = e
                wait = _RETRY_BASE_SEC * (2 ** attempt)
                if status == 429:
                    ra = e.response.headers.get("Retry-After")
                    if ra:
                        wait = max(wait, float(ra))
                logger.warning(
                    f"[FETCH] HTTP {status}"
                    f" ({attempt+1}/{_RETRY_MAX})"
                    f" — {wait:.0f}秒待機"
                )
                time.sleep(wait)
            else:
                raise
    raise last_exc  # type: ignore


def _fetch_all(
    rest_url: str,
    table: str,
    select: str,
    headers: dict,
    extra_params: dict | None = None,
    order: str = "",
) -> list[dict]:
    """
    Supabase REST API からページネーション付きで全件取得する。
    Rangeヘッダーを使い _PAGE_SIZE ずつ取得。
    """
    all_rows: list[dict] = []
    offset = 0

    while True:
        range_start = offset
        range_end = offset + _PAGE_SIZE - 1

        req_headers = {
            **headers,
            "Range": f"{range_start}-{range_end}",
            "Prefer": "count=exact",
        }
        params = {"select": select}
        if order:
            params["order"] = order
        if extra_params:
            params.update(extra_params)

        r = _fetch_with_retry(
            f"{rest_url}/{table}",
            headers=req_headers,
            params=params,
        )
        rows = r.json()
        if not rows:
            break
        all_rows.extend(rows)

        if len(rows) < _PAGE_SIZE:
            break
        offset += _PAGE_SIZE

    return all_rows


# ============================================================
# エラーメッセージ（日本語・初心者向け）
# ============================================================
class UserError(Exception):
    """初心者向けの日本語エラーメッセージを持つ例外"""
    pass


def _friendly_error(e: Exception) -> str:
    """例外から初心者向けの日本語メッセージを生成"""
    msg = str(e)

    if isinstance(e, UserError):
        return msg

    if isinstance(e, requests.ConnectionError):
        return (
            "インターネット接続を確認してください。\n"
            "  Wi-Fi / LANケーブルの接続を確認してから再実行してください。"
        )

    if isinstance(e, requests.Timeout):
        return (
            "サーバーからの応答がありませんでした。\n"
            "  しばらく待ってから再実行してください。"
        )

    if isinstance(e, requests.HTTPError):
        status = getattr(e.response, "status_code", None)
        if status == 401 or status == 403:
            return (
                "データベースへの接続が拒否されました。\n"
                "  .env ファイルの SUPABASE_URL と SUPABASE_ANON_KEY が正しいか確認してください。"
            )
        return (
            f"データベースからエラーが返されました（コード: {status}）。\n"
            "  しばらく待ってから再実行してください。"
        )

    if isinstance(e, PermissionError):
        return (
            "ファイルの保存先に書き込み権限がありません。\n"
            "  OneDriveの同期が完了するまで待ってから再実行してください。"
        )

    if isinstance(e, OSError) and "being used" in msg.lower():
        return (
            "data.xlsx が他のプログラムで開かれています。\n"
            "  data.xlsx を閉じてから再実行してください。"
        )

    return f"予期せぬエラーが発生しました: {msg}"


# ============================================================
# メイン: Supabase financials → data.xlsx
# ============================================================

def generate(
    output_path: str = "data/data.xlsx",
    supabase_url: str = "",
    supabase_key: str = "",
    sqlite_db: str = "",
) -> dict:
    """
    Supabase public.financials から全行を取得し、
    data.xlsx の DATA シートへフラットテーブルとして書き出す。

    列: ticker | period | quarter | sales | gross_profit | operating_profit
        | source | updated_at

    Returns: {"rows": int, "tickers": int, "generated_at": str, "errors": int}
    """
    generated_at = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

    # --- 接続情報の解決 ---
    if not supabase_url or not supabase_key:
        _load_dotenv()
        supabase_url = supabase_url or os.environ.get("SUPABASE_URL", "")
        supabase_key = supabase_key or os.environ.get(
            "SUPABASE_SERVICE_ROLE_KEY", ""
        ) or os.environ.get("SUPABASE_ANON_KEY", "")

    if not supabase_url or not supabase_key:
        raise UserError(
            ".env ファイルが見つからないか、接続情報が未設定です。\n"
            "  プロジェクトフォルダに .env ファイルがあることを確認してください。\n"
            "  .env に SUPABASE_URL と SUPABASE_ANON_KEY / SUPABASE_SERVICE_ROLE_KEY が必要です。"
        )

    rest_url = supabase_url.rstrip("/") + "/rest/v1"
    api_headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
    }

    # --- financials テーブルから全件取得 ---
    logger.info("public.financials を取得中...")
    financials = _fetch_all(
        rest_url,
        "financials",
        "ticker,period,quarter,sales,gross_profit,operating_profit,source,updated_at",
        api_headers,
        order="ticker,period,quarter",
    )
    logger.info(f"  → {len(financials):,} 行取得")

    if not financials:
        raise UserError(
            "financials テーブルにデータがありません。\n"
            "  先に sync_financials.py を実行してください。"
        )

    # --- ticker 正規化 + ソート ---
    # Python 安定ソート3回: 最後にソートした key が最優先
    # 1) quarter DESC
    financials.sort(key=lambda r: _quarter_sort_key(r.get("quarter", "")), reverse=True)
    # 2) period DESC
    financials.sort(key=lambda r: r.get("period", ""), reverse=True)
    # 3) ticker ASC (4桁正規化後)
    financials.sort(key=lambda r: _normalize_ticker(r.get("ticker", "")))
    logger.info("  → ソート完了: ticker ASC → period DESC → quarter DESC")

    # --- ticker統計 ---
    unique_tickers = set()
    for row in financials:
        unique_tickers.add(_normalize_ticker(row.get("ticker", "")))
    logger.info(f"  → {len(unique_tickers):,} tickers")

    # --- data.xlsx 作成 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # A1: 生成日時
    ws.cell(row=1, column=1, value=f"generated_at: {generated_at}")

    # ヘッダー (2行目)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # データ行 (3行目から)
    row_idx = 3
    for rec in financials:
        ticker = _normalize_ticker(rec.get("ticker", ""))
        period = rec.get("period", "")
        quarter = rec.get("quarter", "")
        ws.cell(row=row_idx, column=1, value=ticker)
        ws.cell(row=row_idx, column=2, value=period)
        ws.cell(row=row_idx, column=3, value=quarter)
        ws.cell(row=row_idx, column=4, value=rec.get("sales"))
        ws.cell(row=row_idx, column=5, value=rec.get("gross_profit"))
        ws.cell(row=row_idx, column=6, value=rec.get("operating_profit"))
        ws.cell(row=row_idx, column=7, value=rec.get("source", ""))
        ws.cell(row=row_idx, column=8, value=rec.get("updated_at", ""))
        ws.cell(row=row_idx, column=9, value=_make_recency_key(period, quarter))
        ws.cell(row=row_idx, column=10, value=_make_lookup_key(ticker, period, quarter))
        row_idx += 1

    total = row_idx - 3

    # --- 列幅調整 ---
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 24

    # --- SEGMENT_DATA シート追加 (SQLite から読み取り) ---
    seg_total = 0
    seg_db_path = sqlite_db or _DEFAULT_SQLITE_DB
    if not Path(seg_db_path).is_absolute():
        seg_db_path = str(Path(_PROJECT_ROOT) / seg_db_path)
    logger.info(f"[SEGMENT] SQLite DB: {seg_db_path}")

    if os.path.exists(seg_db_path):
        seg_conn = sqlite3.connect(seg_db_path, timeout=10)
        try:
            _ensure_segment_financials_clean_view(seg_conn)
        finally:
            seg_conn.close()

    seg_data = _read_segment_data(seg_db_path)
    if seg_data:
        seg_ws = wb.create_sheet(SEGMENT_SHEET_NAME)
        seg_ws.cell(row=1, column=1, value=f"generated_at: {generated_at}")
        for col, h in enumerate(SEGMENT_HEADERS, 1):
            cell = seg_ws.cell(row=2, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        for i, row_data in enumerate(seg_data, 3):
            for c, val in enumerate(row_data, 1):
                seg_ws.cell(row=i, column=c, value=val)
        seg_total = len(seg_data)
        seg_ws.column_dimensions["A"].width = 10
        seg_ws.column_dimensions["B"].width = 14
        seg_ws.column_dimensions["C"].width = 8
        seg_ws.column_dimensions["D"].width = 24
        seg_ws.column_dimensions["E"].width = 18
        seg_ws.column_dimensions["F"].width = 18
        seg_ws.column_dimensions["G"].width = 12
        seg_ws.column_dimensions["H"].width = 24
        logger.info(f"[GEN] SEGMENT_DATA シート: {seg_total:,} 行")
    else:
        logger.info("[GEN] SEGMENT_DATA: データなし — シートスキップ")

    # --- extracted_facts 3シート追加 ---
    ef_stats = {"forecast_rows": 0, "monthly_rows": 0, "kpi_rows": 0}
    if os.path.exists(seg_db_path):
        ef_conn = sqlite3.connect(seg_db_path, timeout=10)
        try:
            ef_stats = write_extracted_facts_sheets(wb, ef_conn, generated_at)
        finally:
            ef_conn.close()
        logger.info(
            f"[GEN] FORECAST={ef_stats['forecast_rows']:,} "
            f"MONTHLY={ef_stats['monthly_rows']:,} "
            f"KPI={ef_stats['kpi_rows']:,}"
        )

    # --- 原子的保存（temp → 置換） ---
    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx.tmp", dir=str(output.parent)
    )
    os.close(tmp_fd)

    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, str(output))
    except PermissionError:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise UserError(
            "data.xlsx が他のプログラムで開かれています。\n"
            "  data.xlsx を閉じてから再実行してください。"
        )
    except OSError as e:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        if "being used" in str(e).lower() or "denied" in str(e).lower():
            raise UserError(
                "data.xlsx が他のプログラムで開かれています。\n"
                "  data.xlsx を閉じてから再実行してください。"
            )
        raise

    logger.info(f"[GEN] data.xlsx 生成完了: {total:,} 行 -> {output}")
    logger.info(f"[GEN] === 検証サマリ ===")
    logger.info(f"[GEN]   financials     : {len(financials):,} 行")
    logger.info(f"[GEN]   tickers        : {len(unique_tickers):,}")
    logger.info(f"[GEN]   DATA出力行数   : {total:,} 行")
    logger.info(f"[GEN]   SEGMENT_DATA   : {seg_total:,} 行")
    logger.info(f"[GEN]   生成日時       : {generated_at}")
    logger.info(f"[GEN]   出力先         : {output}")
    return {
        "rows": total,
        "tickers": len(unique_tickers),
        "segment_rows": seg_total,
        "forecast_rows": ef_stats.get("forecast_rows", 0),
        "monthly_rows": ef_stats.get("monthly_rows", 0),
        "kpi_rows": ef_stats.get("kpi_rows", 0),
        "generated_at": generated_at,
        "errors": 0,
    }


# ============================================================
# CLI
# ============================================================

def main():
    import io as _io

    # Windows コンソール UTF-8 対応
    if sys.stdout and hasattr(sys.stdout, "encoding"):
        if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
            sys.stdout = _io.TextIOWrapper(
                sys.stdout.buffer, encoding="utf-8", errors="replace"
            )
    if sys.stderr and hasattr(sys.stderr, "encoding"):
        if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
            sys.stderr = _io.TextIOWrapper(
                sys.stderr.buffer, encoding="utf-8", errors="replace"
            )

    parser = argparse.ArgumentParser(description="Supabase financials → data.xlsx 生成")
    parser.add_argument(
        "--output", "-o", default="data/data.xlsx",
        help="出力パス (default: data/data.xlsx)",
    )
    parser.add_argument(
        "--db", default=_DEFAULT_SQLITE_DB,
        help=f"SQLite DB パス — SEGMENT_DATA用 (default: {_DEFAULT_SQLITE_DB})",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="[%(asctime)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    print()
    print("=" * 50)
    print("  data.xlsx 生成ツール")
    print("  ソース: Supabase public.financials")
    print("=" * 50)
    print()

    try:
        result = generate(args.output, sqlite_db=args.db)

        print("=" * 50)
        print("  ✅ 成功")
        print("=" * 50)
        print(f"  出力ファイル : {args.output}")
        print(f"  DATA行数     : {result['rows']:,} 行")
        print(f"  SEGMENT行数  : {result.get('segment_rows', 0):,} 行")
        print(f"  FORECAST行数 : {result.get('forecast_rows', 0):,} 行")
        print(f"  MONTHLY行数  : {result.get('monthly_rows', 0):,} 行")
        print(f"  KPI行数      : {result.get('kpi_rows', 0):,} 行")
        print(f"  ticker数     : {result['tickers']:,}")
        print(f"  生成日時     : {result['generated_at']}")
        print(f"  エラー       : {result['errors']}")
        print("=" * 50)
        print()
        sys.exit(0)

    except (UserError, requests.ConnectionError, requests.Timeout,
            requests.HTTPError, PermissionError, OSError) as e:
        msg = _friendly_error(e)
        print()
        print("=" * 50)
        print("  ❌ エラーが発生しました")
        print("=" * 50)
        print()
        for line in msg.split("\n"):
            print(f"  {line}")
        print()
        print("=" * 50)
        print()
        sys.exit(1)

    except Exception as e:
        print()
        print("=" * 50)
        print("  ❌ 予期せぬエラー")
        print("=" * 50)
        traceback.print_exc()
        print()
        print("  このメッセージが出た場合は管理者に連絡してください。")
        print("=" * 50)
        print()
        sys.exit(1)


if __name__ == "__main__":
    main()
