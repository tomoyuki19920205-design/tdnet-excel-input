#!/usr/bin/env python3
# ============================================================
# excel_sync.py — DB ↔ Excel 同期ツール
# ============================================================
#
# push: DB→Excel（O,P,S,U,V,Y列 + セグメント列 + 数式列）
# pull: Excel→DB（Z列メモのみ）
# watch: xlwingsでExcelイベント監視（Z列メモ自動保存）
#
# CLI:
#   python -m tools.excel_sync --db data/xbrl.db --excel <path> --mode push
#   python -m tools.excel_sync --db data/xbrl.db --excel <path> --mode pull
#   python -m tools.excel_sync --db data/xbrl.db --excel <path> --mode watch
#
# ============================================================
from __future__ import annotations

import argparse
import logging
import os
import re
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

logger = logging.getLogger("excel_sync")

JST = timezone(timedelta(hours=9))

# ============================================================
# Excel列定義
# ============================================================
# 列名→0-indexed 変換
_COL = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ")}
_COL.update({
    "AA": 26, "AB": 27, "AC": 28, "AD": 29, "AE": 30, "AF": 31,
    "AG": 32, "AH": 33, "AI": 34, "AJ": 35, "AK": 36, "AL": 37,
})

# 列仕様
COL_CODE      = "A"    # 企業コード
COL_FY        = "M"    # 会計年 (R8/3等)
COL_Q         = "N"    # 四半期 (1Q..4Q)
COL_SALES     = "O"    # 累計売上
COL_GP        = "P"    # 累計粗利
COL_GPM       = "Q"    # 粗利率 =P/O（数式）
COL_SGA       = "R"    # 販管費 =P-S（数式）
COL_OP        = "S"    # 累計営業利益
COL_OPM       = "T"    # 営利率 =S/O（数式）
COL_SALES_S   = "U"    # 単体売上
COL_GP_S      = "V"    # 単体粗利
COL_GPM_S     = "W"    # 単体粗利率 =V/U（数式）
COL_SGA_S     = "X"    # 単体販管費 =V-Y（数式）
COL_OP_S      = "Y"    # 単体営業利益
COL_MEMO      = "Z"    # メモ
COL_SEG_START = "AA"   # セグメント開始列

# ============================================================
# R表記パース
# ============================================================

_REIWA_BASE = 2018  # 令和1年 = 2019年

def _parse_fy_label(label: str) -> tuple[str, int] | None:
    """
    'R8/3' → ('2026-03-31', 2026) のように
    Excel M列の会計年ラベルを fiscal_year_end に変換する。
    """
    import calendar
    m = re.match(r"R(\d+)/(\d{1,2})", str(label).strip())
    if not m:
        return None
    ad_year = _REIWA_BASE + int(m.group(1))
    month = int(m.group(2))
    if not (1 <= month <= 12):
        return None
    last_day = calendar.monthrange(ad_year, month)[1]
    return f"{ad_year:04d}-{month:02d}-{last_day:02d}", ad_year


def _parse_quarter_label(label: str) -> int | None:
    """'1Q' → 1, '2Q' → 2, ..., '4Q' → 4"""
    m = re.match(r"(\d)Q", str(label).strip())
    if not m:
        return None
    q = int(m.group(1))
    return q if 1 <= q <= 4 else None


# ============================================================
# DB操作ヘルパー
# ============================================================

def _get_company_id(conn: sqlite3.Connection, ticker: str) -> int | None:
    cur = conn.execute(
        "SELECT company_id FROM companies WHERE ticker_code = ?", (ticker,)
    )
    row = cur.fetchone()
    return row[0] if row else None


def _get_period_id(
    conn: sqlite3.Connection, company_id: int,
    fiscal_year_end: str, quarter: int,
) -> int | None:
    cur = conn.execute(
        "SELECT period_id FROM periods "
        "WHERE company_id = ? AND fiscal_year_end = ? AND quarter = ?",
        (company_id, fiscal_year_end, quarter),
    )
    row = cur.fetchone()
    return row[0] if row else None


def _get_latest_fact(
    conn: sqlite3.Connection, company_id: int, period_id: int,
    metric: str, scope: str,
) -> int | None:
    """v_latest_factsから値を取得（円整数）"""
    cur = conn.execute(
        "SELECT value FROM v_latest_facts "
        "WHERE company_id = ? AND period_id = ? AND metric = ? AND scope = ?",
        (company_id, period_id, metric, scope),
    )
    row = cur.fetchone()
    return row[0] if row else None


def _get_segments(
    conn: sqlite3.Connection, company_id: int, period_id: int,
) -> list[dict]:
    """segment_factsから取得（segment_order順）"""
    cur = conn.execute(
        "SELECT segment_name, segment_order, sales, profit "
        "FROM segment_facts "
        "WHERE company_id = ? AND period_id = ? "
        "ORDER BY segment_order",
        (company_id, period_id),
    )
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _get_memo(
    conn: sqlite3.Connection, company_id: int, period_id: int,
) -> str | None:
    cur = conn.execute(
        "SELECT memo_text FROM quarterly_memos "
        "WHERE company_id = ? AND period_id = ?",
        (company_id, period_id),
    )
    row = cur.fetchone()
    return row[0] if row else None


def _upsert_memo(
    conn: sqlite3.Connection, company_id: int, period_id: int,
    memo_text: str,
) -> str:
    """quarterly_memosをUPSERT。Returns: 'inserted'|'updated'|'no_change'"""
    existing = _get_memo(conn, company_id, period_id)
    now = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

    if existing is None:
        conn.execute(
            "INSERT INTO quarterly_memos (company_id, period_id, memo_text, updated_at) "
            "VALUES (?, ?, ?, ?)",
            (company_id, period_id, memo_text, now),
        )
        return "inserted"

    if existing == memo_text:
        return "no_change"

    conn.execute(
        "UPDATE quarterly_memos SET memo_text = ?, updated_at = ? "
        "WHERE company_id = ? AND period_id = ?",
        (memo_text, now, company_id, period_id),
    )
    return "updated"


# ============================================================
# Excel単位変換（DB: 円整数 → Excel: 百万円）
# ============================================================

def _jpy_to_display(value: int | None, unit: str = "百万円") -> float | None:
    """円整数 → 百万円（Excel表示用）"""
    if value is None:
        return None
    if unit == "百万円":
        return value / 1_000_000
    if unit == "千円":
        return value / 1_000
    return float(value)


# ============================================================
# OneDrive安全書き込み
# ============================================================

def _safe_save_excel(wb, excel_path: str) -> None:
    """
    OneDrive共有ファイルでも壊れにくい保存方式。
    1. 一時ファイルに保存
    2. 成功したら元ファイルと入れ替え (atomic rename)
    ファイルロック時はリトライ。
    """
    excel_dir = os.path.dirname(os.path.abspath(excel_path))
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=excel_dir)
    os.close(fd)

    try:
        wb.save(tmp_path)
        # バックアップ（上書き前の安全コピー）
        bak_path = excel_path + ".bak"
        if os.path.exists(excel_path):
            shutil.copy2(excel_path, bak_path)
        # 入れ替え
        shutil.move(tmp_path, excel_path)
        logger.info(f"[SAVE] OneDrive安全保存完了: {excel_path}")
    except PermissionError:
        # OneDriveロック時: tmpから手動コピーを案内
        logger.error(
            f"[SAVE] ファイルロック中。一時ファイルに保存済み: {tmp_path}\n"
            f"  手動でコピーしてください: copy {tmp_path} {excel_path}"
        )
        raise
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# ============================================================
# Supabase → Excel push（REST API直接、SQLite不要）
# ============================================================

def push_from_supabase(
    excel_path: str,
    sheet_name: str = "PL",
    excel_unit: str = "百万円",
    max_rows: int = 500,
    supabase_url: str = "",
    supabase_key: str = "",
) -> dict:
    """
    Supabase v_latest_facts → Excel へ書き込む。
    SQLite不要。.envから接続情報を読む。

    上書きルール:
      - DBに値があるセルのみ書き込む
      - DBに値がない(NULL)セルは触らない（既存値を保持）
      - 数式列(Q,R,T,W,X)は常に設定する

    Returns: {"rows_scanned": int, "rows_updated": int, "rows_skipped": int}
    """
    import openpyxl
    import requests

    # .env読み込み
    if not supabase_url or not supabase_key:
        _load_dotenv()
        supabase_url = supabase_url or os.environ.get("SUPABASE_URL", "")
        supabase_key = supabase_key or os.environ.get("SUPABASE_ANON_KEY", "")

    if not supabase_url or not supabase_key:
        raise ValueError("SUPABASE_URL / SUPABASE_ANON_KEY が未設定")

    rest_url = supabase_url.rstrip("/") + "/rest/v1"
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
    }

    # 1. Supabaseから全データを一括取得（API呼び出し最小化）
    r = requests.get(
        f"{rest_url}/companies?select=company_id,ticker_code",
        headers=headers,
    )
    r.raise_for_status()
    company_map = {c["ticker_code"]: c["company_id"] for c in r.json()}

    r = requests.get(
        f"{rest_url}/periods?select=period_id,company_id,fiscal_year_end,quarter",
        headers=headers,
    )
    r.raise_for_status()
    period_map = {}
    for p in r.json():
        key = (p["company_id"], p["fiscal_year_end"], p["quarter"])
        period_map[key] = p["period_id"]

    r = requests.get(
        f"{rest_url}/v_latest_facts?select=company_id,period_id,metric,scope,value",
        headers=headers,
    )
    r.raise_for_status()
    facts_map = {}
    for f in r.json():
        key = (f["company_id"], f["period_id"], f["metric"], f["scope"])
        facts_map[key] = f["value"]

    # 2. Excel操作
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    result = {"rows_scanned": 0, "rows_updated": 0, "rows_skipped": 0}

    for row_idx in range(2, max_rows + 2):
        ticker = ws.cell(row=row_idx, column=_COL[COL_CODE] + 1).value
        fy_label = ws.cell(row=row_idx, column=_COL[COL_FY] + 1).value
        q_label = ws.cell(row=row_idx, column=_COL[COL_Q] + 1).value

        if not ticker or not fy_label or not q_label:
            continue

        result["rows_scanned"] += 1
        ticker = str(ticker).strip()

        parsed = _parse_fy_label(fy_label)
        if parsed is None:
            result["rows_skipped"] += 1
            continue
        fiscal_year_end, _ = parsed

        quarter = _parse_quarter_label(q_label)
        if quarter is None:
            result["rows_skipped"] += 1
            continue

        company_id = company_map.get(ticker)
        if company_id is None:
            result["rows_skipped"] += 1
            continue

        period_id = period_map.get((company_id, fiscal_year_end, quarter))
        if period_id is None:
            result["rows_skipped"] += 1
            continue

        r_row = row_idx
        updated = False

        def _fact(metric, scope="CONSOLIDATED"):
            return _jpy_to_display(
                facts_map.get((company_id, period_id, metric, scope)),
                excel_unit,
            )

        # --- 値書き込み（Noneならスキップ=既存値保持）---
        for col, val in [
            (COL_SALES, _fact("NET_SALES")),
            (COL_GP,    _fact("GROSS_PROFIT")),
            (COL_OP,    _fact("OP_INCOME")),
            (COL_SALES_S, _fact("NET_SALES", "NON_CONSOLIDATED")),
            (COL_GP_S,    _fact("GROSS_PROFIT", "NON_CONSOLIDATED")),
            (COL_OP_S,    _fact("OP_INCOME", "NON_CONSOLIDATED")),
        ]:
            if val is not None:
                ws.cell(row=r_row, column=_COL[col] + 1, value=val)
                updated = True

        # --- 数式列（常に設定）---
        ws.cell(row=r_row, column=_COL[COL_GPM] + 1,
                value=f'=IF({COL_SALES}{r_row}=0,"",{COL_GP}{r_row}/{COL_SALES}{r_row})')
        ws.cell(row=r_row, column=_COL[COL_SGA] + 1,
                value=f'={COL_GP}{r_row}-{COL_OP}{r_row}')
        ws.cell(row=r_row, column=_COL[COL_OPM] + 1,
                value=f'=IF({COL_SALES}{r_row}=0,"",{COL_OP}{r_row}/{COL_SALES}{r_row})')
        ws.cell(row=r_row, column=_COL[COL_GPM_S] + 1,
                value=f'=IF({COL_SALES_S}{r_row}=0,"",{COL_GP_S}{r_row}/{COL_SALES_S}{r_row})')
        ws.cell(row=r_row, column=_COL[COL_SGA_S] + 1,
                value=f'={COL_GP_S}{r_row}-{COL_OP_S}{r_row}')

        if updated:
            result["rows_updated"] += 1

    _safe_save_excel(wb, excel_path)

    logger.info(
        f"[PUSH-SUPA] scanned={result['rows_scanned']} "
        f"updated={result['rows_updated']} skipped={result['rows_skipped']}"
    )
    return result


def _load_dotenv():
    """簡易.envパーサー"""
    env_path = Path(_PROJECT_ROOT) / ".env"
    if not env_path.exists():
        return
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, val = line.split("=", 1)
                os.environ.setdefault(key.strip(), val.strip())


# ============================================================
# push モード: DB→Excel (openpyxl + SQLite)
# ============================================================

def push_to_excel(
    db_path: str,
    excel_path: str,
    sheet_name: str = "PL",
    excel_unit: str = "百万円",
    max_rows: int = 500,
) -> dict:
    """
    DB の v_latest_facts + segment_facts → Excel へ書き込む。
    openpyxl で操作（ファイル閉じた状態で実行）。

    Returns: {"rows_scanned": int, "rows_updated": int, "errors": list}
    """
    import openpyxl

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    result = {"rows_scanned": 0, "rows_updated": 0, "errors": []}

    for row_idx in range(2, max_rows + 2):  # 1行目はヘッダー
        ticker = ws.cell(row=row_idx, column=_COL[COL_CODE] + 1).value
        fy_label = ws.cell(row=row_idx, column=_COL[COL_FY] + 1).value
        q_label = ws.cell(row=row_idx, column=_COL[COL_Q] + 1).value

        if not ticker or not fy_label or not q_label:
            continue

        result["rows_scanned"] += 1
        ticker = str(ticker).strip()

        # M列 → fiscal_year_end
        parsed = _parse_fy_label(fy_label)
        if parsed is None:
            continue
        fiscal_year_end, _ = parsed

        # N列 → quarter
        quarter = _parse_quarter_label(q_label)
        if quarter is None:
            continue

        # DB照合
        company_id = _get_company_id(conn, ticker)
        if company_id is None:
            continue

        period_id = _get_period_id(conn, company_id, fiscal_year_end, quarter)
        if period_id is None:
            continue

        r = row_idx  # 行番号

        # 連結値
        sales = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "NET_SALES", "CONSOLIDATED"),
            excel_unit,
        )
        gp = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "GROSS_PROFIT", "CONSOLIDATED"),
            excel_unit,
        )
        op = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "OP_INCOME", "CONSOLIDATED"),
            excel_unit,
        )

        # 単体値
        sales_s = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "NET_SALES", "NON_CONSOLIDATED"),
            excel_unit,
        )
        gp_s = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "GROSS_PROFIT", "NON_CONSOLIDATED"),
            excel_unit,
        )
        op_s = _jpy_to_display(
            _get_latest_fact(conn, company_id, period_id, "OP_INCOME", "NON_CONSOLIDATED"),
            excel_unit,
        )

        updated = False

        # O列: 累計売上
        if sales is not None:
            ws.cell(row=r, column=_COL[COL_SALES] + 1, value=sales)
            updated = True
        # P列: 累計粗利
        if gp is not None:
            ws.cell(row=r, column=_COL[COL_GP] + 1, value=gp)
            updated = True
        # S列: 累計営業利益
        if op is not None:
            ws.cell(row=r, column=_COL[COL_OP] + 1, value=op)
            updated = True

        # 数式列
        ws.cell(row=r, column=_COL[COL_GPM] + 1,
                value=f"=IF({COL_SALES}{r}=0,\"\",{COL_GP}{r}/{COL_SALES}{r})")
        ws.cell(row=r, column=_COL[COL_SGA] + 1,
                value=f"={COL_GP}{r}-{COL_OP}{r}")
        ws.cell(row=r, column=_COL[COL_OPM] + 1,
                value=f"=IF({COL_SALES}{r}=0,\"\",{COL_OP}{r}/{COL_SALES}{r})")

        # 単体値
        if sales_s is not None:
            ws.cell(row=r, column=_COL[COL_SALES_S] + 1, value=sales_s)
            updated = True
        if gp_s is not None:
            ws.cell(row=r, column=_COL[COL_GP_S] + 1, value=gp_s)
            updated = True
        if op_s is not None:
            ws.cell(row=r, column=_COL[COL_OP_S] + 1, value=op_s)
            updated = True

        # 単体数式列
        ws.cell(row=r, column=_COL[COL_GPM_S] + 1,
                value=f"=IF({COL_SALES_S}{r}=0,\"\",{COL_GP_S}{r}/{COL_SALES_S}{r})")
        ws.cell(row=r, column=_COL[COL_SGA_S] + 1,
                value=f"={COL_GP_S}{r}-{COL_OP_S}{r}")

        # セグメント列（AA以降、2列ずつ）
        segments = _get_segments(conn, company_id, period_id)
        seg_col_start = _COL[COL_SEG_START] + 1  # 1-indexed
        for i, seg in enumerate(segments):
            col_sales = seg_col_start + (i * 2)
            col_profit = seg_col_start + (i * 2) + 1
            seg_sales = _jpy_to_display(seg.get("sales"), excel_unit)
            seg_profit = _jpy_to_display(seg.get("profit"), excel_unit)
            if seg_sales is not None:
                ws.cell(row=r, column=col_sales, value=seg_sales)
            if seg_profit is not None:
                ws.cell(row=r, column=col_profit, value=seg_profit)
            updated = True

        if updated:
            result["rows_updated"] += 1

    wb.save(excel_path)
    conn.close()

    logger.info(
        f"[PUSH] 完了: scanned={result['rows_scanned']} "
        f"updated={result['rows_updated']}"
    )
    return result


# ============================================================
# pull モード: Excel→DB (openpyxl, Z列メモのみ)
# ============================================================

def pull_from_excel(
    db_path: str,
    excel_path: str,
    sheet_name: str = "PL",
    max_rows: int = 500,
) -> dict:
    """
    Excel の Z列メモ → DB quarterly_memos へ保存。
    openpyxl で読み取り専用で開く。

    Returns: {"rows_scanned": int, "memos_saved": int, "memos_unchanged": int}
    """
    import openpyxl

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    result = {"rows_scanned": 0, "memos_saved": 0, "memos_unchanged": 0}

    for row_idx in range(2, max_rows + 2):
        row_data = []
        for col_idx in range(1, _COL[COL_MEMO] + 2):  # A〜Z列
            row_data.append(ws.cell(row=row_idx, column=col_idx).value)

        ticker = row_data[_COL[COL_CODE]]
        fy_label = row_data[_COL[COL_FY]]
        q_label = row_data[_COL[COL_Q]]
        memo = row_data[_COL[COL_MEMO]]

        if not ticker or not fy_label or not q_label:
            continue

        result["rows_scanned"] += 1
        ticker = str(ticker).strip()
        memo_text = str(memo).strip() if memo is not None else ""

        if not memo_text:
            continue

        # M列 → fiscal_year_end
        parsed = _parse_fy_label(fy_label)
        if parsed is None:
            continue
        fiscal_year_end, _ = parsed

        # N列 → quarter
        quarter = _parse_quarter_label(q_label)
        if quarter is None:
            continue

        # DB照合
        company_id = _get_company_id(conn, ticker)
        if company_id is None:
            continue

        period_id = _get_period_id(conn, company_id, fiscal_year_end, quarter)
        if period_id is None:
            continue

        status = _upsert_memo(conn, company_id, period_id, memo_text)
        if status in ("inserted", "updated"):
            result["memos_saved"] += 1
        else:
            result["memos_unchanged"] += 1

    conn.commit()
    wb.close()
    conn.close()

    logger.info(
        f"[PULL] 完了: scanned={result['rows_scanned']} "
        f"saved={result['memos_saved']} unchanged={result['memos_unchanged']}"
    )
    return result


# ============================================================
# watch モード: win32com SheetChange イベント駆動
# ============================================================
# セル確定時にそのセルだけ保存。行全体スキャンなし。
# 対象列: Z列（メモ）  ※将来 C〜E 列も追加可能
# ============================================================

# watch対象列（1-indexed）
_WATCH_COLUMNS: set[int] = {
    _COL[COL_MEMO] + 1,     # Z列 = 26
    # 将来拡張: _COL["C"] + 1, _COL["D"] + 1, _COL["E"] + 1,
}


def _handle_cell_change(
    db_path: str, xl_app, sh, target,
) -> None:
    """
    SheetChange イベントハンドラ（単一セル）。
    対象列のセルが確定されたときにDBへ即UPSERTする。
    ダイアログ禁止。成功/失敗はステータスバーのみ。
    """
    try:
        col = target.Column
        row = target.Row

        # 対象列でなければ即return
        if col not in _WATCH_COLUMNS:
            return
        # ヘッダー行は無視
        if row < 2:
            return

        # 変更セルの行から (A, M, N) を読み取り
        ticker = sh.Cells(row, _COL[COL_CODE] + 1).Value
        fy_label = sh.Cells(row, _COL[COL_FY] + 1).Value
        q_label = sh.Cells(row, _COL[COL_Q] + 1).Value

        if not ticker or not fy_label or not q_label:
            return

        ticker = str(ticker).strip()

        # (M, N) → period_id 解決
        parsed = _parse_fy_label(str(fy_label))
        if parsed is None:
            return
        fiscal_year_end, _ = parsed

        quarter = _parse_quarter_label(str(q_label))
        if quarter is None:
            return

        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA foreign_keys = ON")

        company_id = _get_company_id(conn, ticker)
        if company_id is None:
            conn.close()
            return

        period_id = _get_period_id(conn, company_id, fiscal_year_end, quarter)
        if period_id is None:
            conn.close()
            return

        # --- 列に応じた保存 ---
        if col == _COL[COL_MEMO] + 1:
            memo_val = target.Value
            memo_text = str(memo_val).strip() if memo_val is not None else ""
            status = _upsert_memo(conn, company_id, period_id, memo_text)
            conn.commit()
            conn.close()

            if status in ("inserted", "updated"):
                xl_app.StatusBar = (
                    f"[DB保存OK] {ticker} {fy_label} {q_label} "
                    f"メモ{status}"
                )
            else:
                xl_app.StatusBar = (
                    f"[DB] {ticker} {fy_label} {q_label} 変更なし"
                )
            logger.info(
                f"[WATCH] memo {status}: {ticker} {fy_label} {q_label}"
            )
        else:
            conn.close()

    except Exception as e:
        logger.error(f"[WATCH] エラー: {e}")
        try:
            xl_app.StatusBar = f"[DB保存エラー] {e}"
        except Exception:
            pass


def watch_excel(db_path: str, excel_path: str, sheet_name: str = "PL"):
    """
    win32com を使って Excel の SheetChange イベントを待ち受ける。
    セル編集が確定されたタイミングで、そのセルだけDBに保存する。
    行全体スキャンなし。ポーリングなし。
    Ctrl+C で終了。
    """
    import pythoncom
    import win32com.client

    print("=" * 60)
    print("  Excel メモ自動保存 (イベント駆動)")
    print("=" * 60)
    print(f"  DB    : {db_path}")
    print(f"  Excel : {excel_path}")
    print(f"  Sheet : {sheet_name}")
    print()
    print("  対象列を編集するとDBに即保存されます。")
    print(f"  監視列 : Z (メモ)")
    print("  Ctrl+C で終了。")
    print("=" * 60)

    abs_path = os.path.abspath(excel_path)

    # COM初期化
    pythoncom.CoInitialize()

    # 既存のExcelプロセスにアタッチ or 新規起動
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
        print("[INFO] 既存のExcelプロセスにアタッチ")
    except Exception:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = True
        print("[INFO] Excel新規起動")

    # ファイルが開いているか確認
    wb = None
    for i in range(1, xl.Workbooks.Count + 1):
        try:
            if os.path.abspath(xl.Workbooks(i).FullName) == abs_path:
                wb = xl.Workbooks(i)
                break
        except Exception:
            pass

    if wb is None:
        wb = xl.Workbooks.Open(abs_path)
        print(f"[INFO] ファイルを開きました: {abs_path}")

    xl.StatusBar = f"[DB同期] イベント監視中... ({sheet_name})"
    print("[INFO] SheetChange イベント待受開始")

    # --- イベントシンクをセットアップ ---
    # win32com の WithEvents パターン
    class AppEvents:
        """Excel Application イベントシンク"""

        def OnSheetChange(self, Sh, Target):
            """セル編集確定時に呼ばれる (Application.SheetChange)"""
            _handle_cell_change(db_path, xl, Sh, Target)

    # イベントシンクを接続
    events = win32com.client.WithEvents(xl, AppEvents)

    try:
        print("[INFO] Ctrl+C で終了。Excelで自由に編集してください。")
        # メッセージポンプ: COMイベントを待つ
        pythoncom.PumpMessages()
    except KeyboardInterrupt:
        print("\n[INFO] watch 終了")
        xl.StatusBar = ""
    finally:
        pythoncom.CoUninitialize()


# ============================================================
# CLI エントリポイント
# ============================================================

def main():
    import io as _io

    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
        sys.stdout = _io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )

    parser = argparse.ArgumentParser(
        description="DB / Supabase <-> Excel 同期ツール"
    )
    parser.add_argument("--excel", type=str, required=True, help="Excelファイルパス")
    parser.add_argument(
        "--mode", type=str, required=True,
        choices=["push", "push-supabase", "pull", "watch"],
        help=(
            "push=SQLite->Excel, push-supabase=Supabase->Excel, "
            "pull=Excel->DB(メモ), watch=イベント監視"
        ),
    )
    parser.add_argument("--db", type=str, default="data/xbrl.db",
                        help="SQLiteパス (push/pull/watch用, デフォルト: data/xbrl.db)")
    parser.add_argument("--sheet", type=str, default="PL", help="シート名")
    parser.add_argument("--unit", type=str, default="百万円",
                        help="表示単位 (百万円/千円/円)")
    parser.add_argument("--max-rows", type=int, default=500)
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="[%(asctime)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    if args.mode == "push-supabase":
        print("=" * 60)
        print("  Supabase -> Excel (push-supabase)")
        print("=" * 60)
        result = push_from_supabase(
            args.excel, args.sheet, args.unit, args.max_rows,
        )
        print(f"  走査行数  : {result['rows_scanned']}")
        print(f"  更新行数  : {result['rows_updated']}")
        print(f"  スキップ  : {result['rows_skipped']}")

    elif args.mode == "push":
        print("=" * 60)
        print("  SQLite -> Excel (push)")
        print("=" * 60)
        result = push_to_excel(
            args.db, args.excel, args.sheet, args.unit, args.max_rows,
        )
        print(f"  走査行数  : {result['rows_scanned']}")
        print(f"  更新行数  : {result['rows_updated']}")

    elif args.mode == "pull":
        print("=" * 60)
        print("  Excel -> DB (pull)")
        print("=" * 60)
        result = pull_from_excel(
            args.db, args.excel, args.sheet, args.max_rows,
        )
        print(f"  走査行数  : {result['rows_scanned']}")
        print(f"  保存件数  : {result['memos_saved']}")
        print(f"  変更なし  : {result['memos_unchanged']}")

    elif args.mode == "watch":
        watch_excel(args.db, args.excel, args.sheet)

    print()


if __name__ == "__main__":
    main()
