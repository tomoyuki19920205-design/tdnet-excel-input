#!/usr/bin/env python3
# ============================================================
# generate_data_excel.py
# ============================================================
# Supabase (v_latest_facts) → data.xlsx (フラットテーブル)
#
# ダブルクリック実行:
#   run_generate_data.bat
#
# コマンド実行:
#   .\.venv\Scripts\python.exe -m tools.generate_data_excel
#
# ============================================================
from __future__ import annotations

import argparse
import logging
import os
import shutil
import sys
import tempfile
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import requests

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

logger = logging.getLogger("gen_data")

JST = timezone(timedelta(hours=9))

# ============================================================
# 定数
# ============================================================
SHEET_NAME = "DATA"
HEADERS = [
    "ticker",
    "period",
    "quarter",
    "net_sales",
    "gross_profit",
    "operating_profit",
    "source_doc_id",
    "disclosed_at",
    "updated_at",
]

_PAGE_SIZE = 1000  # Supabase 1リクエストあたりの上限

# ============================================================
# R表記変換 (fiscal_year_end -> R8/3)
# ============================================================
_REIWA_BASE = 2018


def _fye_to_label(fye: str) -> str:
    """'2026-03-31' -> 'R8/3'"""
    try:
        parts = fye.split("-")
        year = int(parts[0])
        month = int(parts[1])
        reiwa = year - _REIWA_BASE
        return f"R{reiwa}/{month}"
    except Exception:
        return fye


def _quarter_label(q: int) -> str:
    return f"{q}Q"


# ============================================================
# .env 読み込み
# ============================================================
def _load_dotenv():
    env_path = Path(_PROJECT_ROOT) / ".env"
    if not env_path.exists():
        return False
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, val = line.split("=", 1)
                os.environ.setdefault(key.strip(), val.strip())
    return True


# ============================================================
# Supabase ページネーション付き取得
# ============================================================
def _fetch_all(
    rest_url: str,
    table: str,
    select: str,
    headers: dict,
    extra_params: dict | None = None,
) -> list[dict]:
    """
    Supabase REST API からページネーション付きで全件取得する。
    Rangeヘッダーを使い _PAGE_SIZE ずつ取得。
    """
    all_rows: list[dict] = []
    offset = 0

    while True:
        range_start = offset
        range_end = offset + _PAGE_SIZE - 1

        req_headers = {
            **headers,
            "Range": f"{range_start}-{range_end}",
            "Prefer": "count=exact",
        }
        params = {"select": select}
        if extra_params:
            params.update(extra_params)

        r = requests.get(
            f"{rest_url}/{table}",
            headers=req_headers,
            params=params,
            timeout=30,
        )
        r.raise_for_status()
        rows = r.json()
        if not rows:
            break
        all_rows.extend(rows)

        # 取得件数が _PAGE_SIZE 未満なら最終ページ
        if len(rows) < _PAGE_SIZE:
            break
        offset += _PAGE_SIZE

    return all_rows


# ============================================================
# エラーメッセージ（日本語・初心者向け）
# ============================================================
class UserError(Exception):
    """初心者向けの日本語エラーメッセージを持つ例外"""
    pass


def _friendly_error(e: Exception) -> str:
    """例外から初心者向けの日本語メッセージを生成"""
    msg = str(e)

    if isinstance(e, UserError):
        return msg

    if isinstance(e, requests.ConnectionError):
        return (
            "インターネット接続を確認してください。\n"
            "  Wi-Fi / LANケーブルの接続を確認してから再実行してください。"
        )

    if isinstance(e, requests.Timeout):
        return (
            "サーバーからの応答がありませんでした。\n"
            "  しばらく待ってから再実行してください。"
        )

    if isinstance(e, requests.HTTPError):
        status = getattr(e.response, "status_code", None)
        if status == 401 or status == 403:
            return (
                "データベースへの接続が拒否されました。\n"
                "  .env ファイルの SUPABASE_URL と SUPABASE_ANON_KEY が正しいか確認してください。"
            )
        return (
            f"データベースからエラーが返されました（コード: {status}）。\n"
            "  しばらく待ってから再実行してください。"
        )

    if isinstance(e, PermissionError):
        return (
            "ファイルの保存先に書き込み権限がありません。\n"
            "  OneDriveの同期が完了するまで待ってから再実行してください。"
        )

    if isinstance(e, OSError) and "being used" in msg.lower():
        return (
            "data.xlsx が他のプログラムで開かれています。\n"
            "  data.xlsx を閉じてから再実行してください。"
        )

    return f"予期せぬエラーが発生しました: {msg}"


# ============================================================
# メイン: Supabase -> data.xlsx
# ============================================================

def generate(
    output_path: str = "data/data.xlsx",
    supabase_url: str = "",
    supabase_key: str = "",
) -> dict:
    """
    Supabase から companies / periods / v_latest_facts / disclosures を取得し、
    data.xlsx の DATA シートへフラットテーブルとして書き出す。

    列: ticker | period | quarter | net_sales | gross_profit | operating_profit
        | source_doc_id | disclosed_at | updated_at

    Returns: {"rows": int, "updated_at": str, "errors": int}
    """
    now_str = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

    # --- 接続情報の解決 ---
    if not supabase_url or not supabase_key:
        env_found = _load_dotenv()
        supabase_url = supabase_url or os.environ.get("SUPABASE_URL", "")
        supabase_key = supabase_key or os.environ.get("SUPABASE_ANON_KEY", "")

    if not supabase_url or not supabase_key:
        raise UserError(
            ".env ファイルが見つからないか、接続情報が未設定です。\n"
            "  プロジェクトフォルダに .env ファイルがあることを確認してください。\n"
            "  .env に SUPABASE_URL と SUPABASE_ANON_KEY が必要です。"
        )

    rest_url = supabase_url.rstrip("/") + "/rest/v1"
    api_headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
    }

    # --- Supabase データ取得（ページネーション付き） ---
    logger.info("companies を取得中...")
    companies = _fetch_all(
        rest_url, "companies", "company_id,ticker_code", api_headers
    )
    cid_to_ticker = {c["company_id"]: c["ticker_code"] for c in companies}
    logger.info(f"  → {len(companies)} 社")

    logger.info("periods を取得中...")
    periods = _fetch_all(
        rest_url,
        "periods",
        "period_id,company_id,fiscal_year_end,quarter",
        api_headers,
    )
    pid_to_info: dict[int, dict] = {}
    for p in periods:
        pid_to_info[p["period_id"]] = {
            "company_id": p["company_id"],
            "fye": p["fiscal_year_end"],
            "quarter": p["quarter"],
        }
    logger.info(f"  → {len(periods)} 期間")

    logger.info("v_latest_facts を取得中...")
    facts_raw = _fetch_all(
        rest_url,
        "v_latest_facts",
        "company_id,period_id,disclosure_id,metric,scope,value,disclosed_at,created_at",
        api_headers,
    )
    logger.info(f"  → {len(facts_raw)} 件")

    # facts を (company_id, period_id) -> {metric: value, ...} に集約
    row_data: dict[tuple, dict] = {}
    for f in facts_raw:
        if f.get("scope") != "CONSOLIDATED":
            continue
        key = (f["company_id"], f["period_id"])
        if key not in row_data:
            row_data[key] = {
                "disclosure_id": None,
                "disclosed_at": None,
            }
        row_data[key][f["metric"]] = f["value"]
        # 最新の disclosure_id / disclosed_at を保持
        if f.get("disclosure_id"):
            row_data[key]["disclosure_id"] = f["disclosure_id"]
        if f.get("disclosed_at"):
            existing = row_data[key]["disclosed_at"]
            if existing is None or f["disclosed_at"] > existing:
                row_data[key]["disclosed_at"] = f["disclosed_at"]

    logger.info(f"集約後: {len(row_data)} 行")

    # --- data.xlsx 作成（一時ファイルに書き込み） ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # ヘッダー (A1:I1) — 固定9列
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=h)

    # データ行（ticker + period でソート）
    sorted_keys = sorted(
        row_data.keys(),
        key=lambda k: (
            cid_to_ticker.get(k[0], ""),
            pid_to_info.get(k[1], {}).get("fye", ""),
            pid_to_info.get(k[1], {}).get("quarter", 0),
        ),
    )

    row_idx = 2
    for cid, pid in sorted_keys:
        ticker = cid_to_ticker.get(cid, "")
        info = pid_to_info.get(pid, {})
        fye_label = _fye_to_label(info.get("fye", ""))
        q_label = _quarter_label(info.get("quarter", 0))

        facts = row_data[(cid, pid)]

        ws.cell(row=row_idx, column=1, value=ticker)          # A: ticker
        ws.cell(row=row_idx, column=2, value=fye_label)        # B: period
        ws.cell(row=row_idx, column=3, value=q_label)          # C: quarter
        ws.cell(row=row_idx, column=4, value=facts.get("NET_SALES"))        # D
        ws.cell(row=row_idx, column=5, value=facts.get("GROSS_PROFIT"))     # E
        ws.cell(row=row_idx, column=6, value=facts.get("OP_INCOME"))        # F
        ws.cell(row=row_idx, column=7, value=facts.get("disclosure_id"))    # G: source_doc_id
        ws.cell(row=row_idx, column=8, value=facts.get("disclosed_at"))     # H: disclosed_at
        ws.cell(row=row_idx, column=9, value=now_str)                       # I: updated_at

        row_idx += 1

    total = row_idx - 2

    # --- 原子的保存（temp → 置換） ---
    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    # 一時ファイルに保存
    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx.tmp", dir=str(output.parent)
    )
    os.close(tmp_fd)

    try:
        wb.save(tmp_path)
        # 一時ファイル → 本体に置換（原子的）
        shutil.move(tmp_path, str(output))
    except PermissionError:
        # 一時ファイル清掃
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise UserError(
            "data.xlsx が他のプログラムで開かれています。\n"
            "  data.xlsx を閉じてから再実行してください。"
        )
    except OSError as e:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        if "being used" in str(e).lower() or "denied" in str(e).lower():
            raise UserError(
                "data.xlsx が他のプログラムで開かれています。\n"
                "  data.xlsx を閉じてから再実行してください。"
            )
        raise

    logger.info(f"[GEN] data.xlsx 生成完了: {total} 行 -> {output}")
    logger.info(f"[GEN] === 検証サマリ ===")
    logger.info(f"[GEN]   companies      : {len(companies)} 社")
    logger.info(f"[GEN]   periods        : {len(periods)} 期間")
    logger.info(f"[GEN]   v_latest_facts : {len(facts_raw)} 件 (raw)")
    logger.info(f"[GEN]   集約後行数     : {len(row_data)} 行")
    logger.info(f"[GEN]   出力行数       : {total} 行")
    logger.info(f"[GEN]   出力先         : {output}")
    return {"rows": total, "updated_at": now_str, "errors": 0}


# ============================================================
# CLI
# ============================================================

def main():
    import io as _io

    # Windows コンソール UTF-8 対応
    if sys.stdout and hasattr(sys.stdout, "encoding"):
        if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
            sys.stdout = _io.TextIOWrapper(
                sys.stdout.buffer, encoding="utf-8", errors="replace"
            )
    if sys.stderr and hasattr(sys.stderr, "encoding"):
        if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
            sys.stderr = _io.TextIOWrapper(
                sys.stderr.buffer, encoding="utf-8", errors="replace"
            )

    parser = argparse.ArgumentParser(description="Supabase -> data.xlsx 生成")
    parser.add_argument(
        "--output", "-o", default="data/data.xlsx",
        help="出力パス (default: data/data.xlsx)",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="[%(asctime)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    print()
    print("=" * 50)
    print("  data.xlsx 生成ツール")
    print("=" * 50)
    print()

    try:
        result = generate(args.output)

        print("=" * 50)
        print("  ✅ 成功")
        print("=" * 50)
        print(f"  出力ファイル : {args.output}")
        print(f"  データ行数   : {result['rows']} 行")
        print(f"  更新日時     : {result['updated_at']}")
        print(f"  エラー       : {result['errors']}")
        print("=" * 50)
        print()
        sys.exit(0)

    except (UserError, requests.ConnectionError, requests.Timeout,
            requests.HTTPError, PermissionError, OSError) as e:
        msg = _friendly_error(e)
        print()
        print("=" * 50)
        print("  ❌ エラーが発生しました")
        print("=" * 50)
        print()
        for line in msg.split("\n"):
            print(f"  {line}")
        print()
        print("=" * 50)
        print()
        sys.exit(1)

    except Exception as e:
        print()
        print("=" * 50)
        print("  ❌ 予期せぬエラー")
        print("=" * 50)
        print(f"  {e}")
        print()
        print("  このメッセージが出た場合は管理者に連絡してください。")
        print("=" * 50)
        print()
        sys.exit(1)


if __name__ == "__main__":
    main()
