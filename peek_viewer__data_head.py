from openpyxl import load_workbook

VIEWER = r"C:\Users\takuy\OneDrive\20260303テスト用コピー.xlsx"

wb = load_workbook(VIEWER, data_only=False, read_only=True)

# _DATAの先頭数行だけ見る（ticker列が 2590 か 25900 か）
if "_DATA" not in wb.sheetnames:
    print("FATAL: _DATA sheet not found. sheets=", wb.sheetnames)
    raise SystemExit(1)

ws = wb["_DATA"]

# 1〜10行目を表示
for r in range(1, 11):
    vals = []
    for c in ws.iter_rows(min_row=r, max_row=r, values_only=True):
        vals = list(c)
    print(r, vals[:8])

wb.close()
