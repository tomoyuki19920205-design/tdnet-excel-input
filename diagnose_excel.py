"""
セグメントパース デバッグ診断
parse_excelで実際にセグメントが幾つ取れるかを直接テスト。
Usage: python diagnose_excel.py <excel_path> [sheet_name]
"""
import sys, os
from src.migration.excel_parser import parse_excel, IDX_AA

def main():
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else "PL"

    print(f"=== parse_excel 実行 ===")
    result = parse_excel(path, sheet)

    print(f"Blocks: {len(result.blocks)}")
    total_q = sum(len(b.records) for b in result.blocks)
    total_seg = sum(len(r.segments) for b in result.blocks for r in b.records)
    q_with_seg = sum(1 for b in result.blocks for r in b.records if r.segments)
    print(f"Total Q records: {total_q}")
    print(f"Total segments: {total_seg}")
    print(f"Q records with segments: {q_with_seg}")

    # サンプル出力
    count = 0
    for b in result.blocks:
        for r in b.records:
            if r.segments:
                if count < 10:
                    print(f"\n  企業={b.company_code} Q={r.quarter} FY={r.fiscal_year_end}")
                    for s in r.segments[:3]:
                        print(f"    seg[{s.segment_order}] {s.segment_name}: "
                              f"sales={s.segment_sales}, profit={s.segment_profit}")
                count += 1

    if total_seg == 0:
        print("\n=== セグメント0件 追加診断 ===")
        # parse直後のブロックで row_values の長さをチェック
        import openpyxl, re
        from src.migration.excel_parser import _ZEN2HAN, _val_num

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet]
        max_col = ws.max_column or 19
        print(f"max_col={max_col}, IDX_AA={IDX_AA}")

        checked = 0
        code = None
        in_block = False
        has_hdr = False
        rows_since = 0
        hdr_row_values = None

        for rn, rv in enumerate(ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True
        ), 1):
            a = rv[0] if len(rv) > 0 else None
            s = str(a).strip().translate(_ZEN2HAN) if a else ""
            if re.match(r"^\d{4,5}$", s):
                code, in_block, has_hdr, rows_since, hdr_row_values = s, True, False, 0, None
                continue
            if not in_block: continue
            rows_since += 1
            if rows_since > 150: in_block, code = False, None; continue

            # ヘッダー検知
            if not has_hdr:
                if len(rv) > 17:
                    o = rv[14]; o_s = str(o).strip() if o else ""
                    if "売上" in o_s:
                        aux = sum(1 for i, k in [(15,"粗利"),(16,"粗利率"),(17,"管理費")]
                                  if rv[i] is not None and k in str(rv[i]).strip())
                        if aux >= 2:
                            has_hdr = True
                            hdr_row_values = rv
                continue

            # Q判定
            n = rv[13] if len(rv) > 13 else None
            ns = str(n).strip().translate(_ZEN2HAN).upper() if n else ""
            if ns not in {"1Q","2Q","3Q","4Q","1","2","3","4"}: continue

            # このQ行のタプル長とAA以降の値を確認
            if checked < 15:
                aa_slice = rv[IDX_AA:] if IDX_AA < len(rv) else ()
                has_any_num = any(_val_num(v) is not None for v in aa_slice[:6])
                print(f"\n  row={rn} code={code} Q={ns} tuple_len={len(rv)}")
                print(f"    AA idx={IDX_AA}, aa_slice_len={len(aa_slice)}")
                print(f"    AA~AF raw: {list(aa_slice[:6])}")
                print(f"    AA~AF types: {[type(v).__name__ for v in aa_slice[:6]]}")
                print(f"    AA~AF _val_num: {[_val_num(v) for v in aa_slice[:6]]}")
                print(f"    has_any_num: {has_any_num}")
                print(f"    header AA~AF: {list(hdr_row_values[IDX_AA:IDX_AA+6]) if hdr_row_values and IDX_AA < len(hdr_row_values) else 'N/A'}")
                checked += 1

        wb.close()
        print(f"\n(checked {checked} Q rows)")


if __name__ == "__main__":
    main()
